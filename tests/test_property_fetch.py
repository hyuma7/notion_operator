"""
Notionプロパティ取得テスト

_flatten_notion_page で各プロパティタイプが正しく取得できるか検証する。
対象プロパティ（SoldRecord / PurchaseRecord で使用）:

【SoldRecord】
  - 商品名         : title
  - 売上金         : number
  - 純利益         : formula (number)
  - 仕入れ原価     : formula (number)
  - 販売手数料     : formula (number)
  - 送料           : formula (number)
  - 売却日         : date
  - 仕入れ先名     : rollup > array > formula(string)
  - 販売媒体名     : rollup > array > formula(string)
  - 仕入先カテゴリ : rollup > array > select
  - 販売先カテゴリ : rollup > array > select
  - 作業担当       : people
  - 販売担当者     : rollup > array > formula(string)

【PurchaseRecord】
  - 仕入れ原価     : formula (number)
  - 仕入れ先名     : rollup > array > formula(string)
  - 仕入先カテゴリ : rollup > array > select
  - 作業担当       : people
  - 仕入れ日       : date（帰属月の基準・サーバーサイドフィルタ対象）

EXCLUDE_PROPERTIES (除外対象 - スキーマに存在しない):
  仕入れ手数料、販売手数料、利益率、純利益（直接数値）、
  送料、仕入手数料、販売手数料率、送料計算方法、
  作成日時、仕入れ日、在庫状況、売上金（直接数値）、
  仕入れ先、仕入れ先名（relation）、購入者名、伝票番号、商品ID、画像URL
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest
from brother_ql_proxy.ui.export.service import ExportService
from brother_ql_proxy.ui.export.schemas import SoldRecord, PurchaseRecord


# ─────────────────────────────────────────────
# ヘルパー: モックページ生成
# ─────────────────────────────────────────────

def make_page(props: dict, created_time: str = "2025-06-15T10:00:00.000Z") -> dict:
    """Notion APIが返すページオブジェクトを模倣する"""
    return {
        "id": "page-test-id",
        "created_time": created_time,
        "properties": props,
    }


def title_prop(text: str) -> dict:
    return {
        "type": "title",
        "title": [{"plain_text": text}] if text else [],
    }


def number_prop(value) -> dict:
    return {"type": "number", "number": value}


def formula_number_prop(value) -> dict:
    return {
        "type": "formula",
        "formula": {"type": "number", "number": value},
    }


def formula_string_prop(value: str) -> dict:
    return {
        "type": "formula",
        "formula": {"type": "string", "string": value},
    }


def date_prop(start: str) -> dict:
    return {
        "type": "date",
        "date": {"start": start, "end": None, "time_zone": None},
    }


def people_prop(names: list) -> dict:
    return {
        "type": "people",
        "people": [{"name": n, "id": f"user-{i}"} for i, n in enumerate(names)],
    }


def rollup_formula_string_prop(value: str) -> dict:
    """rollup > array > formula(string) — 仕入れ先名・販売媒体名・販売担当者"""
    return {
        "type": "rollup",
        "rollup": {
            "type": "array",
            "array": [
                {
                    "type": "formula",
                    "formula": {"type": "string", "string": value},
                }
            ],
        },
    }


def rollup_select_prop(value: str) -> dict:
    """rollup > array > select — 仕入先カテゴリ・販売先カテゴリ"""
    return {
        "type": "rollup",
        "rollup": {
            "type": "array",
            "array": [
                {
                    "type": "select",
                    "select": {"name": value, "id": "sel-id", "color": "blue"},
                }
            ],
        },
    }


def rollup_empty_array_prop() -> dict:
    """空のrollup配列"""
    return {
        "type": "rollup",
        "rollup": {"type": "array", "array": []},
    }


def status_prop(name: str) -> dict:
    return {
        "type": "status",
        "status": {"name": name, "id": "st-id", "color": "green"},
    }


# ─────────────────────────────────────────────
# サービスの準備
# ─────────────────────────────────────────────

@pytest.fixture
def service():
    svc = ExportService.__new__(ExportService)
    svc.api_key = "fake"
    svc.database_id = "fake-db"
    svc.notion = None
    svc._cancelled = False
    svc._progress_callback = None
    svc._property_id_map = None
    return svc


# ─────────────────────────────────────────────
# 1. 商品名 (title)
# ─────────────────────────────────────────────

class TestTitleProperty:
    def test_title_正常取得(self, service):
        page = make_page({"商品名": title_prop("テスト商品ABC")})
        row = service._flatten_notion_page(page)
        assert row["商品名"] == "テスト商品ABC"

    def test_title_空文字列(self, service):
        page = make_page({"商品名": title_prop("")})
        row = service._flatten_notion_page(page)
        assert row["商品名"] == ""

    def test_title_リスト空(self, service):
        page = make_page({"商品名": {"type": "title", "title": []}})
        row = service._flatten_notion_page(page)
        assert row["商品名"] == ""


# ─────────────────────────────────────────────
# 2. 売上金 (number)
# ─────────────────────────────────────────────

class TestNumberProperty:
    def test_number_正整数(self, service):
        page = make_page({"売上金": number_prop(15000)})
        row = service._flatten_notion_page(page)
        assert row["売上金"] == 15000

    def test_number_小数(self, service):
        page = make_page({"売上金": number_prop(12345.67)})
        row = service._flatten_notion_page(page)
        assert row["売上金"] == 12345.67

    def test_number_ゼロ(self, service):
        page = make_page({"売上金": number_prop(0)})
        row = service._flatten_notion_page(page)
        assert row["売上金"] == 0

    def test_number_null(self, service):
        page = make_page({"売上金": number_prop(None)})
        row = service._flatten_notion_page(page)
        # None の場合、get("number", 0) → None が返る
        assert row["売上金"] is None


# ─────────────────────────────────────────────
# 3. 純利益 / 仕入れ原価 / 販売手数料 / 送料 (formula > number)
# ─────────────────────────────────────────────

class TestFormulaNumberProperty:
    @pytest.mark.parametrize("prop_name,value", [
        ("純利益", 3000),
        ("仕入れ原価", 8000),
        ("販売手数料", 500),
        ("送料", 200),
    ])
    def test_formula_number_正常(self, service, prop_name, value):
        page = make_page({prop_name: formula_number_prop(value)})
        row = service._flatten_notion_page(page)
        assert row[prop_name] == value

    def test_formula_number_ゼロ(self, service):
        page = make_page({"純利益": formula_number_prop(0)})
        row = service._flatten_notion_page(page)
        assert row["純利益"] == 0

    def test_formula_number_負値(self, service):
        page = make_page({"純利益": formula_number_prop(-500)})
        row = service._flatten_notion_page(page)
        assert row["純利益"] == -500

    def test_formula_string_も取得できる(self, service):
        """formula(string)タイプも正しく取得できる"""
        page = make_page({"テスト": formula_string_prop("hello")})
        row = service._flatten_notion_page(page)
        assert row["テスト"] == "hello"


# ─────────────────────────────────────────────
# 4. 売却日 (date)
# ─────────────────────────────────────────────

class TestDateProperty:
    def test_date_正常取得(self, service):
        page = make_page({"売却日": date_prop("2025-07-10")})
        row = service._flatten_notion_page(page)
        assert row["売却日"] == "2025-07-10"

    def test_date_null(self, service):
        page = make_page({"売却日": {"type": "date", "date": None}})
        row = service._flatten_notion_page(page)
        assert row["売却日"] is None

    def test_date_datetime形式(self, service):
        page = make_page({"売却日": date_prop("2025-07-10T15:30:00.000+09:00")})
        row = service._flatten_notion_page(page)
        assert row["売却日"] == "2025-07-10T15:30:00.000+09:00"


# ─────────────────────────────────────────────
# 5. 仕入れ先名 / 販売媒体名 / 販売担当者 (rollup > array > formula > string)
# ─────────────────────────────────────────────

class TestRollupFormulaStringProperty:
    @pytest.mark.parametrize("prop_name,value", [
        ("仕入れ先名", "RE"),
        ("販売媒体名", "メルカリ1"),
        ("販売担当者", "齊藤光"),
    ])
    def test_rollup_formula_string_正常(self, service, prop_name, value):
        page = make_page({prop_name: rollup_formula_string_prop(value)})
        row = service._flatten_notion_page(page)
        assert row[prop_name] == value

    def test_rollup_formula_string_空配列(self, service):
        page = make_page({"仕入れ先名": rollup_empty_array_prop()})
        row = service._flatten_notion_page(page)
        assert row["仕入れ先名"] == ""

    def test_rollup_formula_string_空文字値(self, service):
        """formulaの値が空文字列の場合は除外され空文字列になる"""
        page = make_page({"仕入れ先名": rollup_formula_string_prop("")})
        row = service._flatten_notion_page(page)
        assert row["仕入れ先名"] == ""

    def test_rollup_url付き名前(self, service):
        """Notion URL付きの値もrollupからそのまま取得される（cleanはschemaで実施）"""
        raw_name = "RE (https://www.notion.so/RE-27e54e6206d8806083e7e48503f5f080?pvs=21)"
        page = make_page({"仕入れ先名": rollup_formula_string_prop(raw_name)})
        row = service._flatten_notion_page(page)
        # _flatten_notion_page は生のまま返す（URLクリーンはSoldRecord/PurchaseRecordで実施）
        assert row["仕入れ先名"] == raw_name


# ─────────────────────────────────────────────
# 6. 仕入先カテゴリ / 販売先カテゴリ (rollup > array > select)
# ─────────────────────────────────────────────

class TestRollupSelectProperty:
    @pytest.mark.parametrize("prop_name,value", [
        ("仕入先カテゴリ", "ネット"),
        ("仕入先カテゴリ", "市場"),
        ("販売先カテゴリ", "業販"),
        ("販売先カテゴリ", "小売"),
    ])
    def test_rollup_select_正常(self, service, prop_name, value):
        page = make_page({prop_name: rollup_select_prop(value)})
        row = service._flatten_notion_page(page)
        assert row[prop_name] == value

    def test_rollup_select_空配列(self, service):
        page = make_page({"仕入先カテゴリ": rollup_empty_array_prop()})
        row = service._flatten_notion_page(page)
        assert row["仕入先カテゴリ"] == ""


# ─────────────────────────────────────────────
# 7. 作業担当 (people)
# ─────────────────────────────────────────────

class TestPeopleProperty:
    def test_people_1名(self, service):
        page = make_page({"作業担当": people_prop(["齊藤光"])})
        row = service._flatten_notion_page(page)
        assert row["作業担当"] == "齊藤光"

    def test_people_複数名(self, service):
        page = make_page({"作業担当": people_prop(["齊藤光", "ゆかり 齊藤"])})
        row = service._flatten_notion_page(page)
        assert row["作業担当"] == "齊藤光, ゆかり 齊藤"

    def test_people_空リスト(self, service):
        page = make_page({"作業担当": {"type": "people", "people": []}})
        row = service._flatten_notion_page(page)
        assert row["作業担当"] == ""

    def test_people_名前なしユーザー(self, service):
        """nameがないpeopleオブジェクトは除外される"""
        page = make_page({
            "作業担当": {
                "type": "people",
                "people": [{"id": "user-no-name"}],  # nameキーなし
            }
        })
        row = service._flatten_notion_page(page)
        assert row["作業担当"] == ""


# ─────────────────────────────────────────────
# 8. Created time (ページメタデータ)
# ─────────────────────────────────────────────

class TestCreatedTime:
    def test_created_time_マッピング(self, service):
        """ページのcreated_timeが"Created time"キーにマッピングされる"""
        page = make_page({}, created_time="2025-06-15T10:00:00.000Z")
        row = service._flatten_notion_page(page)
        assert row["Created time"] == "2025-06-15T10:00:00.000Z"

    def test_created_time_別の日付(self, service):
        page = make_page({}, created_time="2025-11-01T00:00:00.000Z")
        row = service._flatten_notion_page(page)
        assert row["Created time"] == "2025-11-01T00:00:00.000Z"


# ─────────────────────────────────────────────
# 9. SoldRecord Pydanticバリデーション
# ─────────────────────────────────────────────

class TestSoldRecordValidation:
    def _make_flat(self, **overrides):
        base = {
            "商品名": "テスト商品",
            "売上金": 15000,
            "販売利益": 3000,
            "仕入れ原価": 8000,
            "販売手数料": 500,
            "配送料": 200,
            "売却日": "2025-07-10",
            "仕入れ先名": "RE",
            "販売媒体名": "メルカリ1",
            "仕入先カテゴリ": "ネット",
            "販売先カテゴリ": "業販",
            "作業担当": "齊藤光",
            "販売担当者": "齊藤光",
            "Created time": "2025-07-01T00:00:00.000Z",
        }
        base.update(overrides)
        return base

    def test_正常なレコードが作成される(self):
        record = SoldRecord(**self._make_flat())
        assert record.product_name == "テスト商品"
        assert record.sales_amount == 15000.0
        assert record.profit == 3000.0
        assert record.cost_price == 8000.0
        assert record.sales_channel == "メルカリ1"
        assert record.supplier == "RE"
        assert record.sales_channel_category == "業販"
        assert record.supplier_category == "ネット"

    def test_URL付き仕入れ先名がクリーンされる(self):
        flat = self._make_flat(**{
            "仕入れ先名": "RE (https://www.notion.so/RE-abc123?pvs=21)"
        })
        record = SoldRecord(**flat)
        assert record.supplier == "RE"

    def test_URL付き販売媒体名がクリーンされる(self):
        flat = self._make_flat(**{
            "販売媒体名": "メルカリ1 (https://www.notion.so/abc?pvs=21)"
        })
        record = SoldRecord(**flat)
        assert record.sales_channel == "メルカリ1"

    def test_数値空欄は0になる(self):
        flat = self._make_flat(**{"販売利益": None, "販売手数料": ""})
        record = SoldRecord(**flat)
        assert record.profit == 0.0
        assert record.commission == 0.0

    def test_カテゴリ空欄はその他になる(self):
        flat = self._make_flat(**{"仕入先カテゴリ": "", "販売先カテゴリ": None})
        record = SoldRecord(**flat)
        assert record.supplier_category == "その他"
        assert record.sales_channel_category == "その他"

    def test_仕入れ先名空欄は不明になる(self):
        flat = self._make_flat(**{"仕入れ先名": ""})
        record = SoldRecord(**flat)
        assert record.supplier == "不明"

    def test_販売媒体名空欄は不明になる(self):
        flat = self._make_flat(**{"販売媒体名": ""})
        record = SoldRecord(**flat)
        assert record.sales_channel == "不明"


# ─────────────────────────────────────────────
# 10. PurchaseRecord Pydanticバリデーション
# ─────────────────────────────────────────────

class TestPurchaseRecordValidation:
    def _make_flat(self, **overrides):
        base = {
            "仕入れ原価": 8000,
            "仕入れ先名": "REF",
            "仕入先カテゴリ": "市場",
            "作業担当": "齊藤光",
            "仕入れ日": "2025-07-01",
        }
        base.update(overrides)
        return base

    def test_正常なレコードが作成される(self):
        record = PurchaseRecord(**self._make_flat())
        assert record.cost_price == 8000.0
        assert record.supplier == "REF"
        assert record.supplier_category == "市場"
        assert record.assignee == "齊藤光"

    def test_仕入れ日がpurchase_dateにマッピングされる(self):
        record = PurchaseRecord(**self._make_flat())
        assert record.purchase_date is not None
        assert record.purchase_date.year == 2025
        assert record.purchase_date.month == 7

    def test_仕入れ日未入力はNoneになる(self):
        flat = self._make_flat()
        del flat["仕入れ日"]
        record = PurchaseRecord(**flat)
        assert record.purchase_date is None

    def test_URL付き仕入れ先名がクリーンされる(self):
        flat = self._make_flat(**{
            "仕入れ先名": "REF (https://www.notion.so/REF-abc?pvs=21)"
        })
        record = PurchaseRecord(**flat)
        assert record.supplier == "REF"

    def test_仕入れ原価空欄は0になる(self):
        flat = self._make_flat(**{"仕入れ原価": None})
        record = PurchaseRecord(**flat)
        assert record.cost_price == 0.0

    def test_カテゴリ空欄はその他になる(self):
        flat = self._make_flat(**{"仕入先カテゴリ": ""})
        record = PurchaseRecord(**flat)
        assert record.supplier_category == "その他"


# ─────────────────────────────────────────────
# 11. 統合テスト: _flatten_notion_page → SoldRecord
# ─────────────────────────────────────────────

class TestIntegrationFlattenToSoldRecord:
    def test_完全なページからSoldRecordが作成される(self, service):
        """実際のNotion APIレスポンスに近いモックデータで全プロパティを検証"""
        page = make_page(
            props={
                "商品名": title_prop("iPhone 15 Pro"),
                "売上金": number_prop(120000),
                "販売利益": formula_number_prop(25000),
                "仕入れ原価": formula_number_prop(80000),
                "販売手数料": formula_number_prop(6000),
                "配送料": formula_number_prop(500),
                "売却日": date_prop("2025-08-20"),
                "在庫状況": status_prop("売却済み"),  # フィルタ用、スキーマ外
                "仕入れ先名": rollup_formula_string_prop("RE"),
                "販売媒体名": rollup_formula_string_prop("メルカリ1"),
                "仕入先カテゴリ": rollup_select_prop("ネット"),
                "販売先カテゴリ": rollup_select_prop("業販"),
                "作業担当": people_prop(["齊藤光"]),
                "販売担当者": rollup_formula_string_prop("齊藤光"),
            },
            created_time="2025-06-10T09:00:00.000Z",
        )

        row = service._flatten_notion_page(page)

        # スキーマに変換
        record = SoldRecord(**row)

        assert record.product_name == "iPhone 15 Pro"
        assert record.sales_amount == 120000.0
        assert record.profit == 25000.0
        assert record.cost_price == 80000.0
        assert record.commission == 6000.0
        assert record.shipping_cost == 500.0
        from datetime import date
        assert record.sold_date == date(2025, 8, 20)
        assert record.supplier == "RE"
        assert record.sales_channel == "メルカリ1"
        assert record.supplier_category == "ネット"
        assert record.sales_channel_category == "業販"
        assert record.assignee == "齊藤光"
        assert record.sales_assignee == "齊藤光"

    def test_完全なページからPurchaseRecordが作成される(self, service):
        """仕入データ用の完全なページからPurchaseRecordを作成"""
        page = make_page(
            props={
                "商品名": title_prop("Nintendo Switch"),
                "仕入れ原価": formula_number_prop(25000),
                "仕入れ先名": rollup_formula_string_prop("REF"),
                "仕入先カテゴリ": rollup_select_prop("市場"),
                "作業担当": people_prop(["ゆかり 齊藤"]),
                "仕入れ日": date_prop("2025-09-05"),  # 帰属月の基準
                "在庫状況": status_prop("在庫中"),  # フィルタ用、スキーマ外
            },
            created_time="2025-06-01T12:00:00.000Z",  # 仕入れ日と別日でも帰属月は仕入れ日ベース
        )

        row = service._flatten_notion_page(page)
        record = PurchaseRecord(**row)

        assert record.cost_price == 25000.0
        assert record.supplier == "REF"
        assert record.supplier_category == "市場"
        assert record.assignee == "ゆかり 齊藤"
        assert record.purchase_date is not None
        assert record.purchase_date.year == 2025
        assert record.purchase_date.month == 9


# ─────────────────────────────────────────────
# 12. エッジケース
# ─────────────────────────────────────────────

class TestEdgeCases:
    def test_プロパティが存在しないキーはrowに含まれない(self, service):
        """存在しないプロパティはflatデータに含まれないことを確認"""
        page = make_page({"商品名": title_prop("テスト")})
        row = service._flatten_notion_page(page)
        assert "純利益" not in row
        assert "仕入れ先名" not in row

    def test_id_と_Created_timeは常に存在する(self, service):
        """id と Created time は必ずマッピングされる"""
        page = make_page({}, created_time="2025-01-01T00:00:00.000Z")
        row = service._flatten_notion_page(page)
        assert "id" in row
        assert "Created time" in row
        assert row["id"] == "page-test-id"

    def test_rollup_number型(self, service):
        """rollup(number)タイプも正しく取得できる"""
        page = make_page({
            "数値ロールアップ": {
                "type": "rollup",
                "rollup": {"type": "number", "number": 42.5},
            }
        })
        row = service._flatten_notion_page(page)
        assert row["数値ロールアップ"] == 42.5

    def test_複数formulaアイテムのrollup(self, service):
        """rollup配列に複数formulaアイテムがある場合はカンマ結合"""
        page = make_page({
            "テスト": {
                "type": "rollup",
                "rollup": {
                    "type": "array",
                    "array": [
                        {"type": "formula", "formula": {"type": "string", "string": "A"}},
                        {"type": "formula", "formula": {"type": "string", "string": "B"}},
                    ],
                },
            }
        })
        row = service._flatten_notion_page(page)
        assert row["テスト"] == "A, B"

    def test_status_プロパティ取得(self, service):
        """在庫状況などのstatusプロパティも取得できる"""
        page = make_page({"在庫状況": status_prop("売却済み")})
        row = service._flatten_notion_page(page)
        assert row["在庫状況"] == "売却済み"


# ─────────────────────────────────────────────
# 13. fetch_purchase_data のサーバーサイド date フィルタ
# ─────────────────────────────────────────────

class TestFetchPurchaseDataServerSideFilter:
    def test_仕入れ日のdateフィルタでクエリを組み立てる(self, service):
        service.notion = object()  # truthy でありさえすればよい
        captured = {}

        def fake_query(query_params, filter_properties=None):
            captured["query_params"] = query_params
            captured["filter_properties"] = filter_properties
            return {"results": [], "has_more": False, "next_cursor": None}

        service._query_with_retry = fake_query
        service._get_property_ids = lambda names: []  # ID解決失敗 → 絞り込みなし

        service.fetch_purchase_data("2025-06-01", "2026-06-01")

        assert captured["query_params"]["filter"] == {
            "and": [
                {"property": "仕入れ日", "date": {"on_or_after": "2025-06-01"}},
                {"property": "仕入れ日", "date": {"before": "2026-06-01"}},
            ]
        }
        # ID解決失敗時は filter_properties なしで続行
        assert captured["filter_properties"] is None

    def test_filter_propertiesが解決したIDで渡される(self, service):
        service.notion = object()
        captured = {}

        def fake_query(query_params, filter_properties=None):
            captured["filter_properties"] = filter_properties
            return {"results": [], "has_more": False, "next_cursor": None}

        service._query_with_retry = fake_query
        service._get_property_ids = lambda names: ["id-cost", "id-date"]

        service.fetch_purchase_data("2025-06-01", "2026-06-01")

        assert captured["filter_properties"] == ["id-cost", "id-date"]


# ─────────────────────────────────────────────
# 14. _get_property_ids: 名前→ID解決 とフェイルセーフ
# ─────────────────────────────────────────────

class TestGetPropertyIds:
    def test_名前からIDに解決される(self, service):
        class FakeNotion:
            def request(self, path, method):
                return {
                    "properties": {
                        "仕入れ原価": {"id": "cost-id"},
                        "仕入れ日": {"id": "date-id"},
                        "仕入れ先名": {"id": "sup-id"},
                    }
                }

        service.notion = FakeNotion()
        service._property_id_map = None
        ids = service._get_property_ids(["仕入れ原価", "仕入れ日"])
        assert ids == ["cost-id", "date-id"]

    def test_未知のプロパティは無視される(self, service):
        class FakeNotion:
            def request(self, path, method):
                return {"properties": {"仕入れ原価": {"id": "cost-id"}}}

        service.notion = FakeNotion()
        service._property_id_map = None
        ids = service._get_property_ids(["仕入れ原価", "存在しないプロパティ"])
        assert ids == ["cost-id"]

    def test_解決失敗時は空リストで続行(self, service):
        class FailNotion:
            def request(self, path, method):
                raise RuntimeError("スキーマ取得失敗")

        service.notion = FailNotion()
        service._property_id_map = None
        ids = service._get_property_ids(["仕入れ原価"])
        assert ids == []


# ─────────────────────────────────────────────
# 15. process_pivot_data: フォールバック削除
# ─────────────────────────────────────────────

def _make_sold(month: str, **overrides) -> SoldRecord:
    base = {
        "商品名": "商品",
        "売上金": 10000,
        "販売利益": 2000,
        "仕入れ原価": 6000,
        "販売手数料": 0,
        "配送料": 0,
        "売却日": "2025-07-10",
        "仕入れ先名": "RE",
        "販売媒体名": "メルカリ1",
        "仕入先カテゴリ": "ネット",
        "販売先カテゴリ": "小売",
        "作業担当": "齊藤光",
        "販売担当者": "齊藤光",
    }
    base.update(overrides)
    rec = SoldRecord(**base)
    rec.sold_year_month = month
    return rec


def _make_purchase(month: str, cost: float, supplier: str = "RE") -> PurchaseRecord:
    rec = PurchaseRecord(**{
        "仕入れ原価": cost,
        "仕入れ先名": supplier,
        "仕入先カテゴリ": "ネット",
        "作業担当": "齊藤光",
        "仕入れ日": "2025-07-05",
    })
    rec.purchase_year_month = month
    return rec


class TestPivotFallbackRemoved:
    def test_仕入0件_売上ありでpivot_purchaseは空(self, service):
        months = ["2025年7月"]
        sold = _make_sold("2025年7月")
        result = service.process_pivot_data([sold], [], months)
        assert result["pivot_purchase"].empty


# ─────────────────────────────────────────────
# 16. generate_excel: 全体合算から「仕入高」行を廃止
#     （仕入は「企業別仕入高」セクションで集計する）
# ─────────────────────────────────────────────

class TestSummaryPurchaseRow:
    def _labels(self, path):
        """A列ラベルの集合を返す"""
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        labels = set()
        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                labels.add(row[0])
        return labels

    def test_全体合算に仕入高行が無い(self, service, tmp_path):
        """全体合算から仕入高行は廃止。企業別仕入高セクションは存続する"""
        months = ["2025年7月", "2025年8月"]
        sold = _make_sold("2025年7月")
        purchases = [
            _make_purchase("2025年7月", 5000),
            _make_purchase("2025年7月", 3000),
        ]
        data = service.process_pivot_data([sold], purchases, months)
        out = tmp_path / "summary.xlsx"
        service.generate_excel(str(out), data, months)

        labels = self._labels(out)
        # 全体合算の「仕入高」行は無い
        assert "仕入高" not in labels
        # 全体合算ブロックそのものは存在（売上ベースの項目）
        assert "全体合算" in labels
        assert "売上" in labels
        # 企業別仕入高セクションは存続
        assert "企業別仕入高" in labels

    def test_売上0件なら全体合算は描画されず仕入高セクションのみ(self, service, tmp_path):
        """売上0件だと全体合算ブロックは描画されないが、企業別仕入高は出力される"""
        months = ["2025年7月"]
        purchases = [_make_purchase("2025年7月", 4000)]
        data = service.process_pivot_data([], purchases, months)
        out = tmp_path / "summary_only_purchase.xlsx"
        service.generate_excel(str(out), data, months)

        labels = self._labels(out)
        assert "全体合算" not in labels
        assert "仕入高" not in labels
        assert "企業別仕入高" in labels


if __name__ == "__main__":
    import subprocess
    result = subprocess.run(
        ["python", "-m", "pytest", __file__, "-v", "--tb=short"],
        cwd=os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
    )
    sys.exit(result.returncode)
