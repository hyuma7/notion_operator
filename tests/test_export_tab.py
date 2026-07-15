"""
Excel出力タブのテスト
"""

import unittest
import pandas as pd
import sys
import os
import tempfile
from datetime import datetime
from unittest.mock import patch

# プロジェクトのルートディレクトリをパスに追加
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from brother_ql_proxy.ui.export.ui import (
    ExportTab,
    _build_daily_excel_filename,
)
from brother_ql_proxy.ui.export.service import ExportService
from brother_ql_proxy.ui.export.schemas import (
    SoldRecord, PurchaseRecord, DailySoldRecord, DailyPurchaseRecord,
)
from brother_ql_proxy.ui.file_save import (
    set_saved_file_status,
    should_save_directly,
    unique_file_path,
)


class TestExportTab(unittest.TestCase):
    """ExportTabクラスのテストケース"""

    def setUp(self):
        """各テストの前に実行される初期化処理"""
        # テスト用のダミーデータを作成
        self.test_data = {
            'ID': ['PDT-1', 'PDT-2', 'PDT-3', 'PDT-4'],
            '仕入れ先': ['SA (https://example.com)', 'RE (https://example.com)', 'SA (https://example.com)', ''],
            '販売媒体': ['メルカリ1 (https://example.com)', 'Yahoo! (https://example.com)', 'メルカリ1 (https://example.com)', 'Yahoo! (https://example.com)'],
            '売却日': ['2025-09-15', '2025-10-20', '2025-11-10', '2025-09-25'],
            '売上金': ['￥30,000', '￥20,000', '￥15,000', '￥25,000'],
            '仕入れ原価': ['10000', '8000', '6000', '9000'],
            '販売手数料': ['3000', '2000', '1500', '2500'],
            '送料': ['￥2,200', '￥1,800', '￥1,200', '￥2,000'],
            '純利益': ['14800', '8200', '6300', '11500'],
            '在庫状況': ['売却済み', '売却済み', '売却済み', '売却済み']
        }
        self.df = pd.DataFrame(self.test_data)

    def test_clean_currency(self):
        """通貨クリーニング機能のテスト"""
        # ExportTabのcreate_pivot_sectionsメソッド内のclean_currency関数をテスト
        test_values = [
            ('￥1,000', 1000.0),
            ('2000', 2000.0),
            ('', 0.0),
            (None, 0.0),
            (1500, 1500.0),
            ('￥10,000', 10000.0)
        ]

        for input_val, expected in test_values:
            with self.subTest(input_val=input_val):
                # clean_currency関数のロジックを再現
                if pd.isna(input_val) or input_val == "":
                    result = 0
                elif isinstance(input_val, (int, float)):
                    result = float(input_val)
                else:
                    cleaned = str(input_val).replace('￥', '').replace(',', '').strip()
                    try:
                        result = float(cleaned)
                    except:
                        result = 0

                self.assertEqual(result, expected, f"Failed for input: {input_val}")

    def test_clean_company_name(self):
        """企業名クリーニング機能のテスト"""
        test_cases = [
            ('SA (https://www.notion.so/SA-123?pvs=21)', 'SA'),
            ('メルカリ1 (https://www.notion.so/メルカリ1-456?pvs=21)', 'メルカリ1'),
            ('RE', 'RE'),
            ('', '不明'),
            (None, '不明')
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                import re
                # clean_company_name関数のロジックを再現
                if pd.isna(input_val) or input_val == "":
                    result = "不明"
                else:
                    if '(https://' in str(input_val):
                        result = re.sub(r'\s*\(https://.*?\)', '', str(input_val))
                    else:
                        result = str(input_val)
                    result = result.strip()

                self.assertEqual(result, expected, f"Failed for input: {input_val}")

    def test_extract_year_month(self):
        """年月抽出機能のテスト"""
        test_cases = [
            ('2025-09-15', '2025年9月'),
            ('2025-10-01', '2025年10月'),
            ('2025-11-30', '2025年11月'),
            ('', None),
            (None, None)
        ]

        for input_val, expected in test_cases:
            with self.subTest(input_val=input_val):
                # extract_year_month関数のロジックを再現
                if pd.isna(input_val) or input_val == "":
                    result = None
                else:
                    try:
                        date_obj = pd.to_datetime(input_val)
                        result = f"{date_obj.year}年{date_obj.month}月"
                    except:
                        result = None

                self.assertEqual(result, expected, f"Failed for input: {input_val}")

    def test_month_list_generation(self):
        """12ヶ月の月リスト生成のテスト"""
        start_year = 2025
        start_month = 6

        # 月のリストを生成
        months = []
        current_year = start_year
        current_month = start_month

        for i in range(12):
            months.append(f"{current_year}年{current_month}月")
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

        # 検証
        self.assertEqual(len(months), 12, "月リストは12要素である必要があります")
        self.assertEqual(months[0], "2025年6月", "最初の月が正しくありません")
        self.assertEqual(months[6], "2025年12月", "7番目の月が正しくありません")
        self.assertEqual(months[7], "2026年1月", "8番目の月が正しくありません")
        self.assertEqual(months[11], "2026年5月", "最後の月が正しくありません")

    def test_date_range_calculation(self):
        """期間計算のテスト"""
        test_cases = [
            (2025, 6, "2025-06-01", "2026-06-01"),
            (2025, 1, "2025-01-01", "2026-01-01"),
            (2025, 12, "2025-12-01", "2026-12-01"),
            (2024, 7, "2024-07-01", "2025-07-01")
        ]

        for start_year, start_month, expected_start, expected_end in test_cases:
            with self.subTest(year=start_year, month=start_month):
                # 開始日
                start_date = f"{start_year}-{start_month:02d}-01"

                # 終了日計算
                end_month = start_month + 12
                end_year = start_year
                if end_month > 12:
                    end_year = start_year + 1
                    end_month = end_month - 12

                end_date = f"{end_year}-{end_month:02d}-01"

                self.assertEqual(start_date, expected_start, "開始日が正しくありません")
                self.assertEqual(end_date, expected_end, "終了日が正しくありません")

    def test_pivot_data_structure(self):
        """ピボットデータの構造テスト"""
        # テストデータから簡易的なピボットテーブルを作成
        df_clean = self.df.copy()

        # 日付から年月を抽出
        df_clean['売却年月'] = pd.to_datetime(df_clean['売却日']).apply(
            lambda x: f"{x.year}年{x.month}月"
        )

        # 数値変換
        df_clean['売上金_数値'] = df_clean['売上金'].apply(
            lambda x: float(str(x).replace('￥', '').replace(',', '')) if pd.notna(x) else 0
        )

        # 企業名クリーニング
        import re
        df_clean['仕入れ先_clean'] = df_clean['仕入れ先'].apply(
            lambda x: re.sub(r'\s*\(https://.*?\)', '', str(x)).strip() if pd.notna(x) and x != '' else '不明'
        )

        # ピボットテーブル作成
        pivot = df_clean.pivot_table(
            values='売上金_数値',
            index='仕入れ先_clean',
            columns='売却年月',
            aggfunc='sum',
            fill_value=0
        )

        # 検証
        self.assertGreater(len(pivot), 0, "ピボットテーブルが空です")
        self.assertIn('SA', pivot.index, "SAが仕入先に含まれていません")

    def test_summary_calculation(self):
        """全体合算の計算テスト"""
        # テストデータで全体合算を計算
        df_clean = self.df.copy()

        # 数値変換
        df_clean['売上金_数値'] = df_clean['売上金'].apply(
            lambda x: float(str(x).replace('￥', '').replace(',', '')) if pd.notna(x) else 0
        )
        df_clean['仕入れ原価_数値'] = df_clean['仕入れ原価'].apply(
            lambda x: float(str(x).replace('￥', '').replace(',', '')) if pd.notna(x) else 0
        )
        df_clean['販売手数料_数値'] = df_clean['販売手数料'].apply(
            lambda x: float(str(x).replace('￥', '').replace(',', '')) if pd.notna(x) else 0
        )
        df_clean['送料_数値'] = df_clean['送料'].apply(
            lambda x: float(str(x).replace('￥', '').replace(',', '')) if pd.notna(x) else 0
        )

        # 合計計算
        total_sales = df_clean['売上金_数値'].sum()
        total_cost = df_clean['仕入れ原価_数値'].sum()
        gross_profit = total_sales - total_cost
        total_fees = df_clean['販売手数料_数値'].sum()
        total_shipping = df_clean['送料_数値'].sum()
        net_profit = gross_profit - total_fees - total_shipping

        # 検証
        self.assertEqual(total_sales, 90000, "売上合計が正しくありません")
        self.assertEqual(total_cost, 33000, "原価合計が正しくありません")
        self.assertEqual(gross_profit, 57000, "粗利が正しくありません")
        self.assertGreater(net_profit, 0, "販売利益がマイナスです")


class TestDataIntegration(unittest.TestCase):
    """データ統合のテストケース"""

    def test_company_list_integration(self):
        """仕入先と販売先の統合テスト"""
        # テスト用の企業リスト
        suppliers = ['SA', 'RE', 'ITF']
        retailers = ['メルカリ1', 'Yahoo!', 'SA']

        # 統合
        all_companies = set()
        all_companies.update(suppliers)
        all_companies.update(retailers)
        all_companies.discard('不明')
        all_companies = sorted(list(all_companies))

        # 検証
        self.assertIn('SA', all_companies, "SAが含まれていません")
        self.assertIn('メルカリ1', all_companies, "メルカリ1が含まれていません")
        self.assertNotIn('不明', all_companies, "不明が除外されていません")
        self.assertEqual(len(all_companies), 5, "企業数が正しくありません")


class TestDailyExcelExport(unittest.TestCase):
    """日別Excel保存のテストケース"""

    def test_build_daily_excel_filename(self):
        same_day = datetime(2026, 6, 1)
        self.assertEqual(
            _build_daily_excel_filename(same_day, same_day),
            "日別売上_2026-06-01.xlsx",
        )

        self.assertEqual(
            _build_daily_excel_filename(datetime(2026, 6, 1), datetime(2026, 6, 30)),
            "日別売上_2026-06-01_2026-06-30.xlsx",
        )

    def test_unique_file_path_adds_suffix(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            open(os.path.join(tmpdir, "日別売上_2026-06-01.xlsx"), "w").close()
            open(os.path.join(tmpdir, "日別売上_2026-06-01_2.xlsx"), "w").close()

            self.assertEqual(
                unique_file_path(tmpdir, "日別売上_2026-06-01.xlsx"),
                os.path.join(tmpdir, "日別売上_2026-06-01_3.xlsx"),
            )

    def test_saved_file_status_opens_folder_on_click(self):
        class FakeText:
            def __init__(self):
                self.value = ""
                self.color = None
                self.tooltip = None
                self.on_tap = None
                self.spans = None

        class FakePage:
            def __init__(self):
                self.snack_bar = None
                self.update_count = 0

            def update(self):
                self.update_count += 1

        text = FakeText()
        page = FakePage()

        with patch("brother_ql_proxy.ui.file_save.open_containing_folder") as open_folder:
            set_saved_file_status(text, page, "/tmp/report.xlsx")
            text.spans[0].on_click(None)

        self.assertIn("ファイルを保存しました。", text.value)
        open_folder.assert_called_once_with("/tmp/report.xlsx")
        self.assertIsNone(text.on_tap)

    def test_should_save_directly_by_platform(self):
        class FakePage:
            def __init__(self, web=False):
                self.web = web

        with patch("brother_ql_proxy.ui.file_save.platform.system", return_value="Darwin"):
            self.assertTrue(should_save_directly(FakePage(web=False)))
            self.assertFalse(should_save_directly(FakePage(web=True)))

        with patch("brother_ql_proxy.ui.file_save.platform.system", return_value="Windows"):
            self.assertFalse(should_save_directly(FakePage(web=False)))

    def test_export_daily_excel_saves_directly_on_macos(self):
        class FakeText:
            def __init__(self):
                self.value = ""
                self.color = None
                self.tooltip = None
                self.on_tap = None
                self.spans = None

        class FakePage:
            def __init__(self):
                self.web = False
                self.overlay = []
                self.snack_bar = None
                self.update_count = 0

            def update(self):
                self.update_count += 1

        class FakeService:
            def __init__(self):
                self.calls = []

            def generate_daily_excel(self, path, sales, purchases):
                self.calls.append((path, sales, purchases))
                with open(path, "wb") as f:
                    f.write(b"excel")

        page = FakePage()
        service = FakeService()
        tab = ExportTab.__new__(ExportTab)
        tab.page = page
        tab.service = service
        tab.daily_start_date = datetime(2026, 6, 1)
        tab.daily_end_date = datetime(2026, 6, 1)
        tab.daily_sales_cache = [{"id": "sale-1"}]
        tab.daily_purchases_cache = []
        tab.daily_result_text = FakeText()

        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("brother_ql_proxy.ui.file_save.platform.system", return_value="Darwin"):
                with patch("brother_ql_proxy.ui.file_save.os.path.expanduser", return_value=tmpdir):
                    tab.export_daily_excel(None)

            expected_path = os.path.join(tmpdir, "日別売上_2026-06-01.xlsx")
            self.assertEqual(service.calls, [(expected_path, [{"id": "sale-1"}], [])])
            self.assertTrue(os.path.exists(expected_path))
            self.assertEqual(page.overlay, [])
            self.assertIsNotNone(page.snack_bar)
            self.assertIn("ファイルを保存しました。", tab.daily_result_text.value)
            self.assertIsNotNone(tab.daily_result_text.spans)


class TestGrossProfitSection(unittest.TestCase):
    """全体合算の数式化・仕入高行の廃止・粗利セクション非出力のテスト"""

    def _make_sold(self, month_date, channel, category, sales, cost, gross, profit,
                   commission=0, shipping=0, assignee="担当A"):
        rec = SoldRecord(**{
            "商品名": "テスト商品",
            "売上金": sales,
            "販売利益": profit,
            "粗利": gross,
            "仕入れ原価": cost,
            "販売手数料": commission,
            "配送料": shipping,
            "売却日": month_date,
            "販売媒体名": channel,
            "販売先カテゴリ": category,
            "販売担当者": assignee,
        })
        dt = pd.to_datetime(month_date)
        rec.sold_year_month = f"{dt.year}年{dt.month}月"
        return rec

    def _build_data(self):
        months = ["2026年6月", "2026年7月"]
        sales = [
            # 市場・業販
            self._make_sold("2026-06-10", "市場A", "市場", 30000, 10000, 20000, 17000,
                            commission=2000, shipping=1000),
            self._make_sold("2026-07-05", "業販B", "業販", 20000, 8000, 12000, 10500,
                            commission=1000, shipping=500),
            # 小売り
            self._make_sold("2026-06-20", "メルカリ", "小売", 15000, 6000, 9000, 7500,
                            commission=1200, shipping=300),
        ]
        purchases = [
            PurchaseRecord(**{
                "仕入れ原価": 12000,
                "仕入れ先名": "仕入先X",
                "仕入先カテゴリ": "市場",
                "仕入れ日": "2026-06-01",
            })
        ]
        purchases[0].purchase_year_month = "2026年6月"

        service = ExportService(api_key="", database_id="")
        data = service.process_pivot_data(sales, purchases, months)
        return service, data, months

    def test_process_pivot_includes_gross_pivots(self):
        """process_pivot_data は粗利ピボットを計算し続ける（将来の再有効化用に返り値は残す）"""
        _, data, _ = self._build_data()
        self.assertIn('pivot_gross_wholesale', data)
        self.assertIn('pivot_gross_retail', data)
        self.assertIn('category_gross_wholesale', data)
        self.assertFalse(data['pivot_gross_wholesale'].empty)
        self.assertFalse(data['pivot_gross_retail'].empty)

    def _label_rows(self, path):
        """生成 Excel のラベル(A列) → 行番号のマップを返す"""
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb.active
        label_rows = {}
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and row[0]:
                label_rows.setdefault(row[0], row_idx)
        return ws, label_rows

    def test_gross_sections_not_output(self):
        """企業別粗利セクションはシートに出力されない（計算は残すが出力は見送り）"""
        service, data, months = self._build_data()
        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "summary.xlsx")
            service.generate_excel(out, data, months)
            _, label_rows = self._label_rows(out)

            self.assertNotIn("企業別粗利(市場・業販)", label_rows)
            self.assertNotIn("企業別粗利(小売り)", label_rows)

    def test_summary_has_no_purchase_row(self):
        """全体合算に「仕入高」行が無い（企業別仕入高セクションは存在する）"""
        service, data, months = self._build_data()
        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "summary.xlsx")
            service.generate_excel(out, data, months)
            _, label_rows = self._label_rows(out)

            self.assertNotIn("仕入高", label_rows)
            self.assertIn("企業別仕入高", label_rows)

    def test_summary_formulas_in_excel(self):
        """全体合算の粗利・販売利益・計列が Excel 数式になっている"""
        service, data, months = self._build_data()
        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "summary.xlsx")
            service.generate_excel(out, data, months)
            ws, label_rows = self._label_rows(out)

            # 全体合算の「粗利」行が数式（売上-原価）になっている
            gross_row = label_rows["粗利"]
            gross_cell = ws.cell(row=gross_row, column=2).value
            self.assertTrue(str(gross_cell).startswith("="), "粗利セルが数式でない")
            self.assertIn("-", str(gross_cell))

            # 全体合算の「販売利益」行が数式（粗利-手数料-送料）になっている
            profit_row = label_rows["販売利益"]
            profit_cell = ws.cell(row=profit_row, column=2).value
            self.assertTrue(str(profit_cell).startswith("="), "販売利益セルが数式でない")
            self.assertEqual(str(profit_cell).count("-"), 2, "販売利益は3項の引き算になる")

            # 「計」列（最右列 = len(months)+2）が SUM 数式
            total_cell = ws.cell(row=gross_row, column=len(months) + 2).value
            self.assertTrue(str(total_cell).startswith("=SUM("), "計列がSUM数式でない")


class TestRelationDoubleLinkNormalization(unittest.TestCase):
    """relation 二重リンクで rollup が複数値を ", " join した場合の先頭値採用テスト"""

    def test_sold_supplier_takes_first_value(self):
        """SoldRecord の仕入れ先名が "RE, REO" のとき先頭値 "RE" を採用する"""
        rec = SoldRecord(**{"商品名": "ドラム式洗濯機", "仕入れ先名": "RE, REO"})
        self.assertEqual(rec.supplier, "RE")

    def test_sold_sales_channel_takes_first_value(self):
        """SoldRecord の販売媒体名が複数値のとき先頭値を採用する"""
        rec = SoldRecord(**{"商品名": "商品", "販売媒体名": "メルカリ1, メルカリ2"})
        self.assertEqual(rec.sales_channel, "メルカリ1")

    def test_sold_category_takes_first_value(self):
        """SoldRecord のカテゴリが "ネット, ネット" のとき "ネット" を採用する"""
        rec = SoldRecord(**{"商品名": "商品", "仕入先カテゴリ": "ネット, ネット"})
        self.assertEqual(rec.supplier_category, "ネット")

    def test_purchase_supplier_takes_first_value(self):
        """PurchaseRecord の仕入れ先名が "RE, REO" のとき先頭値 "RE" を採用する"""
        rec = PurchaseRecord(**{"仕入れ先名": "RE, REO", "仕入先カテゴリ": "ネット, ネット"})
        self.assertEqual(rec.supplier, "RE")
        self.assertEqual(rec.supplier_category, "ネット")

    def test_url_and_comma_combined(self):
        """URL 除去後にカンマが残っても先頭値を採用する"""
        rec = SoldRecord(**{
            "商品名": "商品",
            "仕入れ先名": "RE (https://example.com), REO (https://example.com)",
        })
        self.assertEqual(rec.supplier, "RE")

    def test_daily_sold_supplier_and_sales_channel_take_first_value(self):
        """DailySoldRecord の仕入れ先名・販売媒体名が複数値のとき先頭値を採用する"""
        rec = DailySoldRecord(**{
            "商品名": "ドラム式洗濯機",
            "仕入れ先名": "RE, REO",
            "販売媒体名": "メルカリ1, メルカリ2",
        })
        self.assertEqual(rec.supplier, "RE")
        self.assertEqual(rec.sales_channel, "メルカリ1")

    def test_daily_sold_model_number_and_maker_take_first_value(self):
        """DailySoldRecord の型番名・メーカーが複数値のとき先頭値を採用する"""
        rec = DailySoldRecord(**{
            "商品名": "ドラム式洗濯機",
            "型番名": "AA-1, AA-2",
            "メーカー": "パナソニック, シャープ",
        })
        self.assertEqual(rec.model_number, "AA-1")
        self.assertEqual(rec.maker, "パナソニック")

    def test_daily_purchase_supplier_and_category_take_first_value(self):
        """DailyPurchaseRecord の仕入れ先名・仕入先カテゴリが複数値のとき先頭値を採用する"""
        rec = DailyPurchaseRecord(**{
            "仕入れ先名": "RE, REO",
            "仕入先カテゴリ": "ネット, ネット",
        })
        self.assertEqual(rec.supplier, "RE")
        self.assertEqual(rec.supplier_category, "ネット")

    def test_daily_purchase_model_number_maker_category_size_take_first_value(self):
        """DailyPurchaseRecord の型番名・メーカー・カテゴリー・サイズが複数値のとき先頭値を採用する"""
        rec = DailyPurchaseRecord(**{
            "型番名": "AA-1, AA-2",
            "メーカー": "パナソニック, シャープ",
            "カテゴリー": "家電, 家具",
            "サイズ": "大, 小",
        })
        self.assertEqual(rec.model_number, "AA-1")
        self.assertEqual(rec.maker, "パナソニック")
        self.assertEqual(rec.category, "家電")
        self.assertEqual(rec.size, "大")


if __name__ == '__main__':
    # テストを実行
    unittest.main(verbosity=2)
