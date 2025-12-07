"""
テストデータからExcelファイルを生成するスクリプト
"""

import sys
import os
import pandas as pd
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from collections import OrderedDict


def load_test_data():
    """テストCSVデータを読み込む"""
    csv_path = os.path.join(os.path.dirname(__file__), 'data', 'sample_notion_data.csv')
    df = pd.read_csv(csv_path)
    return df


def export_to_excel(sections, output_path, assignee_company_mapping=None, summary_data=None):
    """ピボットセクションをExcelファイルに出力

    Args:
        sections: セクション名 → DataFrameの辞書
        output_path: 出力先パス
        assignee_company_mapping: 担当者名 → 企業リストの辞書（オプション）
        summary_data: 全体合算データのDataFrame（オプション）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "財務集計"

    current_row = 1

    # 色の定義
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 灰色
    total_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # 薄橙色

    # 担当者ごとの色パターン（6色）
    assignee_colors = [
        PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # 薄青
        PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid"),  # 薄ピンク
        PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),  # 薄緑
        PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),  # 薄オレンジ
        PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),  # 薄紫
        PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),  # 薄黄色
    ]

    # 全体合算セクションを最初に書き込み
    if summary_data is not None and not summary_data.empty:
        # セクションタイトル
        ws.cell(row=current_row, column=1, value="全体合算")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1

        # ヘッダー行（月の列） - 灰色背景
        ws.cell(row=current_row, column=1, value="内訳")
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        for col_idx, month in enumerate(summary_data.columns, start=2):
            ws.cell(row=current_row, column=col_idx, value=month)
            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=col_idx).fill = header_fill

        # 「計」列 - 灰色背景
        ws.cell(row=current_row, column=len(summary_data.columns) + 2, value="計")
        ws.cell(row=current_row, column=len(summary_data.columns) + 2).font = Font(bold=True)
        ws.cell(row=current_row, column=len(summary_data.columns) + 2).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=len(summary_data.columns) + 2).fill = header_fill
        current_row += 1

        # データ行
        for item_name in summary_data.index:
            ws.cell(row=current_row, column=1, value=item_name)
            ws.cell(row=current_row, column=1).fill = total_fill
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            row_total = 0
            for col_idx, month in enumerate(summary_data.columns, start=2):
                value = summary_data.loc[item_name, month]
                ws.cell(row=current_row, column=col_idx, value=value)
                ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                ws.cell(row=current_row, column=col_idx).fill = total_fill
                ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                row_total += value

            # 計列
            ws.cell(row=current_row, column=len(summary_data.columns) + 2, value=row_total)
            ws.cell(row=current_row, column=len(summary_data.columns) + 2).number_format = '#,##0'
            ws.cell(row=current_row, column=len(summary_data.columns) + 2).fill = total_fill
            ws.cell(row=current_row, column=len(summary_data.columns) + 2).font = Font(bold=True)
            current_row += 1

        current_row += 1  # 全体合算と各セクションの間に空行

    # 各セクションを順番に書き込み
    for section_name, pivot_df in sections.items():
        # セクションタイトル
        ws.cell(row=current_row, column=1, value=section_name)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1

        # ヘッダー行（月の列） - 灰色背景
        ws.cell(row=current_row, column=1, value="")  # 左上は空欄
        ws.cell(row=current_row, column=1).fill = header_fill
        for col_idx, month in enumerate(pivot_df.columns, start=2):
            ws.cell(row=current_row, column=col_idx, value=month)
            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=col_idx).fill = header_fill

        # 「計」列 - 灰色背景
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value="計")
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = header_fill
        current_row += 1

        # 担当者ごとにグループ化して表示
        if assignee_company_mapping:
            assignee_idx = 0
            for assignee, companies in sorted(assignee_company_mapping.items()):
                # 担当者の色を選択（循環）
                assignee_fill = assignee_colors[assignee_idx % len(assignee_colors)]
                assignee_idx += 1

                # その担当者の企業データを表示 + 小計の計算
                assignee_subtotal_by_month = {month: 0 for month in pivot_df.columns}
                assignee_grand_total = 0

                for company in companies:
                    if company in pivot_df.index:
                        ws.cell(row=current_row, column=1, value=company)
                        row_total = 0
                        for col_idx, month in enumerate(pivot_df.columns, start=2):
                            value = pivot_df.loc[company, month]
                            ws.cell(row=current_row, column=col_idx, value=value)
                            ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                            row_total += value
                            assignee_subtotal_by_month[month] += value

                        # 計列
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=row_total)
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                        assignee_grand_total += row_total
                        current_row += 1

                # 担当者の小計行（担当者名+粗利計）
                ws.cell(row=current_row, column=1, value=f"{assignee}粗利計")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=1).fill = assignee_fill
                for col_idx, month in enumerate(pivot_df.columns, start=2):
                    subtotal_value = assignee_subtotal_by_month[month]
                    ws.cell(row=current_row, column=col_idx, value=subtotal_value)
                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                    ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                    ws.cell(row=current_row, column=col_idx).fill = assignee_fill

                # 小計の計列
                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=assignee_grand_total)
                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = assignee_fill
                current_row += 1
        else:
            # 担当者マッピングがない場合は従来通り企業を表示
            for company in pivot_df.index:
                ws.cell(row=current_row, column=1, value=company)
                row_total = 0
                for col_idx, month in enumerate(pivot_df.columns, start=2):
                    value = pivot_df.loc[company, month]
                    ws.cell(row=current_row, column=col_idx, value=value)
                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                    row_total += value

                # 計列
                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=row_total)
                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                current_row += 1

        # 合計行（薄橙色背景）
        ws.cell(row=current_row, column=1, value="合計")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = total_fill
        for col_idx, month in enumerate(pivot_df.columns, start=2):
            col_total = pivot_df[month].sum()
            ws.cell(row=current_row, column=col_idx, value=col_total)
            ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=current_row, column=col_idx).fill = total_fill

        # 合計の計列
        grand_total = pivot_df.values.sum()
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=grand_total)
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = total_fill
        current_row += 1

        current_row += 1  # セクション間に空行

    # 列幅調整
    ws.column_dimensions['A'].width = 20
    num_cols = len(sections[list(sections.keys())[0]].columns) + 2
    for col_idx in range(2, num_cols + 1):
        if col_idx <= 26:
            col_letter = chr(64 + col_idx)
        else:
            first_letter = chr(64 + (col_idx - 1) // 26)
            second_letter = chr(65 + (col_idx - 1) % 26)
            col_letter = first_letter + second_letter
        ws.column_dimensions[col_letter].width = 12

    # ファイルを保存
    wb.save(output_path)


def create_test_excel():
    """テストデータからExcelファイルを生成"""
    print("テストデータを読み込んでいます...")
    df = load_test_data()

    print(f"データ件数: {len(df)}件")

    # テストデータをDataFrameに変換
    print("データをパースしています...")

    # データクリーニング関数
    def clean_currency(value):
        if pd.isna(value) or value == "":
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        value = str(value).replace('￥', '').replace(',', '').strip()
        try:
            return float(value)
        except:
            return 0

    def clean_company_name(value):
        if pd.isna(value) or value == "":
            return "不明"
        if isinstance(value, str) and '(https://' in value:
            value = re.sub(r'\s*\(https://.*?\)', '', value)
        return value.strip()

    # データ変換
    df_clean = df.copy()
    df_clean['仕入れ先'] = df_clean['仕入れ先'].apply(clean_company_name)
    df_clean['販売媒体'] = df_clean['販売媒体'].apply(clean_company_name)
    df_clean['売上金'] = df_clean['売上金'].apply(clean_currency)
    df_clean['仕入れ原価'] = df_clean['仕入れ原価'].apply(clean_currency)
    df_clean['販売手数料'] = df_clean['販売手数料'].apply(clean_currency)
    df_clean['送料'] = df_clean['送料'].apply(clean_currency)
    df_clean['純利益'] = df_clean['純利益'].apply(clean_currency)

    # 売却年月を抽出（売却済みのみ）
    df_clean['売却年月'] = pd.to_datetime(df_clean['売却日'], errors='coerce').apply(
        lambda x: f"{x.year}年{x.month}月" if pd.notna(x) else None
    )

    # 登録年月を抽出（Created time代わり、仕入高計算用）
    df_clean['登録年月'] = pd.to_datetime(df_clean['登録日'], errors='coerce').apply(
        lambda x: f"{x.year}年{x.month}月" if pd.notna(x) else None
    )

    print("ピボットテーブルを作成しています...")

    # 開始年月（2025年6月）
    start_year = 2025
    start_month = 6

    # 12ヶ月分の月リストを生成
    months = []
    current_year = start_year
    current_month = start_month

    for i in range(12):
        months.append(f"{current_year}年{current_month}月")
        current_month += 1
        if current_month > 12:
            current_month = 1
            current_year += 1

    # 売却済みデータのみをフィルタ（売上・利益計算用）
    df_sold = df_clean[df_clean['在庫状況'] == '売却済み'].copy()

    # ピボットテーブルを作成（5つ）
    pivot_list = []

    # 1. 企業別売上（業販）- 仕入先 × 売上金（売却済みのみ）
    pivot1 = df_sold.pivot_table(
        values='売上金',
        index='仕入れ先',
        columns='売却年月',
        aggfunc='sum',
        fill_value=0
    )

    # 2. 企業別販売利益（業販）- 仕入先 × 純利益（売却済みのみ）
    pivot2 = df_sold.pivot_table(
        values='純利益',
        index='仕入れ先',
        columns='売却年月',
        aggfunc='sum',
        fill_value=0
    )

    # 3. 企業別売上（小売）- 販売媒体 × 売上金（売却済みのみ）
    pivot3 = df_sold.pivot_table(
        values='売上金',
        index='販売媒体',
        columns='売却年月',
        aggfunc='sum',
        fill_value=0
    )

    # 4. 企業別販売利益（小売）- 販売媒体 × 純利益（売却済みのみ）
    pivot4 = df_sold.pivot_table(
        values='純利益',
        index='販売媒体',
        columns='売却年月',
        aggfunc='sum',
        fill_value=0
    )

    # 5. 企業別仕入高 - 仕入先 × 仕入れ原価（全データ、登録年月基準）
    pivot5 = df_clean.pivot_table(
        values='仕入れ原価',
        index='仕入れ先',
        columns='登録年月',
        aggfunc='sum',
        fill_value=0
    )

    pivot_list = [pivot1, pivot2, pivot3, pivot4, pivot5]

    # 存在しない月も0で表示
    for i, pivot in enumerate(pivot_list):
        for month in months:
            if month not in pivot.columns:
                pivot[month] = 0
        pivot_list[i] = pivot[months]

    # Excelに出力
    print("Excelファイルを作成しています...")

    # 5つのセクション構成
    sections = {
        '企業別売上（業販）': pivot_list[0],      # 仕入先 × 売上金
        '企業別販売利益（業販）': pivot_list[1],  # 仕入先 × 純利益
        '企業別売上（小売）': pivot_list[2],      # 販売媒体 × 売上金
        '企業別販売利益（小売）': pivot_list[3],  # 販売媒体 × 純利益
        '企業別仕入高': pivot_list[4]             # 仕入先 × 仕入れ原価（全データ）
    }

    # 全体合算データを作成
    print("全体合算データを作成しています...")
    summary_rows = {}

    # 月ごとに集計
    for month in months:
        month_data = df_sold[df_sold['売却年月'] == month]
        summary_rows.setdefault('売上', {})[month] = month_data['売上金'].sum()
        summary_rows.setdefault('原価', {})[month] = month_data['仕入れ原価'].sum()
        summary_rows.setdefault('販売手数料', {})[month] = month_data['販売手数料'].sum()
        summary_rows.setdefault('送料', {})[month] = month_data['送料'].sum()
        summary_rows.setdefault('販売利益', {})[month] = month_data['純利益'].sum()
        # 粗利 = 売上 - 原価
        summary_rows.setdefault('粗利', {})[month] = summary_rows['売上'][month] - summary_rows['原価'][month]

    # DataFrameに変換（行の順序を指定）
    summary_data = pd.DataFrame(summary_rows).T
    summary_data = summary_data.reindex(['売上', '原価', '粗利', '販売手数料', '送料', '販売利益'])
    summary_data = summary_data[months]  # 月の順序を保証

    # 担当者→企業リストのマッピングを作成
    assignee_company_mapping = {}

    # 仕入先の担当者グループ化
    for _, row in df_clean[['仕入れ先', '作業担当']].drop_duplicates().iterrows():
        supplier = row['仕入れ先']
        assignee = row['作業担当']
        if pd.notna(supplier) and supplier != '不明' and pd.notna(assignee):
            if assignee not in assignee_company_mapping:
                assignee_company_mapping[assignee] = []
            if supplier not in assignee_company_mapping[assignee]:
                assignee_company_mapping[assignee].append(supplier)

    # 販売媒体の担当者グループ化
    for _, row in df_clean[['販売媒体', '作業担当']].drop_duplicates().iterrows():
        channel = row['販売媒体']
        assignee = row['作業担当']
        if pd.notna(channel) and channel != '不明' and pd.notna(assignee):
            if assignee not in assignee_company_mapping:
                assignee_company_mapping[assignee] = []
            if channel not in assignee_company_mapping[assignee]:
                assignee_company_mapping[assignee].append(channel)

    # ファイル名を生成（タイムスタンプ付き）
    end_month = start_month + 11
    end_year = start_year
    if end_month > 12:
        end_year = start_year + 1
        end_month = end_month - 12

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"財務集計_{start_year}年{start_month}月-{end_year}年{end_month}月_テスト_{timestamp}.xlsx"

    output_path = os.path.join(os.path.dirname(__file__), 'data', filename)

    # Excelファイルを出力
    export_to_excel(sections, output_path, assignee_company_mapping, summary_data)

    # 統計用の企業数を計算
    all_companies = set()
    all_companies.update(pivot_list[0].index)  # 仕入先（業販）
    all_companies.update(pivot_list[2].index)  # 販売媒体（小売）
    all_companies.discard('不明')

    print(f"[OK] Excelファイルを作成しました: {output_path}")
    print(f"\n統計情報:")
    print(f"  総売上（売却済み）: {df_sold['売上金'].sum():,.0f}円")
    print(f"  総原価（売却済み）: {df_sold['仕入れ原価'].sum():,.0f}円")
    print(f"  総利益（売却済み）: {df_sold['純利益'].sum():,.0f}円")
    print(f"  総仕入原価（全在庫）: {df_clean['仕入れ原価'].sum():,.0f}円")
    print(f"  企業数: {len(all_companies)}社")
    print(f"  売却済み件数: {len(df_sold)}件")
    print(f"  全データ件数: {len(df_clean)}件")
    print(f"  期間: {start_year}年{start_month}月-{end_year}年{end_month}月")


if __name__ == '__main__':
    create_test_excel()
