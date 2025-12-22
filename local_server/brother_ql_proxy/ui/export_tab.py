"""
Excel出力タブコンポーネント
Notionから売却済みデータを取得してExcel出力

データベース情報:
- データベース名: 商品一覧
- データベースID: 1d254e6206d881bb9e88d2e7ffb90444
- 在庫状況プロパティID: qFjR (status型)
- 売却日プロパティID: ep:H (date型)
"""

import flet as ft
import pandas as pd
from datetime import datetime
from notion_client import Client
import os
from io import BytesIO


class ExportTab:
    def __init__(self, proxy, page: ft.Page):
        self.proxy = proxy
        self.page = page
        self.df = None
        self.data_type = None  # 'yearly' または 'monthly'
        self.fetched_year = None
        self.fetched_month = None
        self.pivot_df = None  # ピボット集計用データ
        self.pivot_start_year_value = None  # ピボット集計の開始年
        self.pivot_start_month_value = None  # ピボット集計の開始月

        # 財務集計用の開始年月選択
        from datetime import datetime
        default_date = datetime(2025, 6, 1)
        self.pivot_start_date = default_date  # 選択された開始日を保存

        # DatePickerの作成
        self.pivot_date_picker = ft.DatePicker(
            value=default_date,
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2030, 12, 31),
            on_change=self.on_pivot_date_change,
            cancel_text="キャンセル",
            confirm_text="選択",
        )

        # 日付選択ボタン
        self.pivot_date_button = ft.ElevatedButton(
            text=f"開始月: {default_date.year}年{default_date.month}月",
            icon=ft.Icons.CALENDAR_MONTH,
            on_click=lambda _: self.page.open(self.pivot_date_picker)
        )

        # ピボット集計用ボタン
        self.fetch_pivot_btn = ft.ElevatedButton(
            "財務集計データを取得",
            icon=ft.Icons.ASSESSMENT,
            on_click=self.fetch_pivot_data
        )

        self.export_pivot_btn = ft.ElevatedButton(
            "財務集計Excelを保存",
            icon=ft.Icons.TABLE_VIEW,
            on_click=self.export_pivot_excel,
            disabled=True
        )

        # 結果表示
        self.result_text = ft.Text("", size=14)

        # プログレスバー
        self.progress = ft.ProgressBar(visible=False)

        # 日別売上用の変数
        self.daily_df = None  # 日別売上データ
        self.daily_purchase_df = None  # 日別仕入データ
        self.daily_start_date = None  # 日別売上の開始日
        self.daily_end_date = None  # 日別売上の終了日

        # 日別売上用の日付選択
        from datetime import timedelta
        today = datetime.now()

        self.daily_start_date = today  # 開始日
        self.daily_end_date = today  # 終了日

        # 開始日DatePicker
        self.daily_start_date_picker_dialog = ft.DatePicker(
            value=today,
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2030, 12, 31),
            on_change=self.on_daily_start_date_change,
            cancel_text="キャンセル",
            confirm_text="選択",
        )

        # 終了日DatePicker
        self.daily_end_date_picker_dialog = ft.DatePicker(
            value=today,
            first_date=datetime(2020, 1, 1),
            last_date=datetime(2030, 12, 31),
            on_change=self.on_daily_end_date_change,
            cancel_text="キャンセル",
            confirm_text="選択",
        )

        # 開始日選択ボタン
        self.daily_start_date_button = ft.ElevatedButton(
            text=f"開始日: {today.strftime('%Y-%m-%d')}",
            icon=ft.Icons.CALENDAR_TODAY,
            on_click=lambda _: self.page.open(self.daily_start_date_picker_dialog)
        )

        # 終了日選択ボタン
        self.daily_end_date_button = ft.ElevatedButton(
            text=f"終了日: {today.strftime('%Y-%m-%d')}",
            icon=ft.Icons.CALENDAR_TODAY,
            on_click=lambda _: self.page.open(self.daily_end_date_picker_dialog)
        )

        # 日別売上用ボタン
        self.fetch_daily_btn = ft.ElevatedButton(
            "日別売上データを取得",
            icon=ft.Icons.CALENDAR_TODAY,
            on_click=self.fetch_daily_data
        )

        self.export_daily_btn = ft.ElevatedButton(
            "日別売上Excelを保存",
            icon=ft.Icons.TABLE_CHART,
            on_click=self.export_daily_excel,
            disabled=True
        )

        # 日別売上用の結果表示
        self.daily_result_text = ft.Text("", size=14)

    def on_pivot_date_change(self, e):
        """財務集計の開始月選択時のイベントハンドラー"""
        if e.control.value:
            self.pivot_start_date = e.control.value
            # ボタンのテキストを更新
            self.pivot_date_button.text = f"開始月: {self.pivot_start_date.year}年{self.pivot_start_date.month}月"
            self.page.update()

    def on_daily_start_date_change(self, e):
        """日別売上の開始日選択時のイベントハンドラー"""
        if e.control.value:
            self.daily_start_date = e.control.value
            # ボタンのテキストを更新
            self.daily_start_date_button.text = f"開始日: {self.daily_start_date.strftime('%Y-%m-%d')}"
            self.page.update()

    def on_daily_end_date_change(self, e):
        """日別売上の終了日選択時のイベントハンドラー"""
        if e.control.value:
            self.daily_end_date = e.control.value
            # ボタンのテキストを更新
            self.daily_end_date_button.text = f"終了日: {self.daily_end_date.strftime('%Y-%m-%d')}"
            self.page.update()

    def fetch_pivot_data(self, e):
        """財務集計データを取得（選択した開始月から12ヶ月分）"""
        api_key = self.proxy.config.get("notion_api_key", "")
        database_id = self.proxy.config.get("notion_database_id", "")

        if not api_key or not database_id:
            self.show_snackbar("設定タブでNotion API KeyとDatabase IDを設定してください", ft.Colors.RED)
            return

        self.progress.visible = True
        self.result_text.value = "財務集計データを取得中..."
        self.page.update()

        try:
            notion = Client(auth=api_key)

            # 選択された開始年月を取得
            start_year = self.pivot_start_date.year
            start_month = self.pivot_start_date.month

            # 開始日と終了日を計算（12ヶ月分）
            start_date = f"{start_year}-{start_month:02d}-01"

            # 12ヶ月後の年月を計算
            end_month = start_month + 12
            end_year = start_year
            if end_month > 12:
                end_year = start_year + 1
                end_month = end_month - 12

            end_date = f"{end_year}-{end_month:02d}-01"

            # 売却済みデータを取得（ページネーション対応）
            all_sold_results = []
            has_more = True
            start_cursor = None

            while has_more:
                query_params = {
                    "database_id": database_id,
                    "filter": {
                        "and": [
                            {
                                "property": "在庫状況",
                                "status": {
                                    "equals": "売却済み"
                                }
                            },
                            {
                                "property": "売却日",
                                "date": {
                                    "on_or_after": start_date
                                }
                            },
                            {
                                "property": "売却日",
                                "date": {
                                    "before": end_date
                                }
                            }
                        ]
                    },
                    "page_size": 100
                }

                if start_cursor:
                    query_params["start_cursor"] = start_cursor

                results = notion.databases.query(**query_params)
                all_sold_results.extend(results["results"])
                has_more = results.get("has_more", False)
                start_cursor = results.get("next_cursor")

            # 全結果を辞書形式に変換
            combined_results = {
                "results": all_sold_results
            }

            # データフレームに変換（売却済みデータ）
            self.pivot_df = self.parse_notion_results(combined_results)

            # 企業別仕入高用：全データを取得（在庫状況不問、Created time基準）
            # ※仕入高は在庫状況に関係なく、登録された全商品の仕入原価を集計
            # Created timeはシステムプロパティのためフィルタできないので、全件取得後にフィルタ
            all_purchase_results = []
            has_more = True
            start_cursor = None

            while has_more:
                query_params = {
                    "database_id": database_id,
                    "page_size": 100
                }

                if start_cursor:
                    query_params["start_cursor"] = start_cursor

                results = notion.databases.query(**query_params)
                all_purchase_results.extend(results["results"])
                has_more = results.get("has_more", False)
                start_cursor = results.get("next_cursor")

            # 仕入データをDataFrameに変換
            purchase_combined_results = {
                "results": all_purchase_results
            }
            self.pivot_purchase_df = self.parse_notion_results(purchase_combined_results)

            # Created timeでフィルタリング（取得後）
            if not self.pivot_purchase_df.empty and 'Created time' in self.pivot_purchase_df.columns:
                # Created timeを日付型に変換（UTC）
                self.pivot_purchase_df['Created time'] = pd.to_datetime(
                    self.pivot_purchase_df['Created time'],
                    errors='coerce',
                    utc=True
                )
                # 期間でフィルタ（UTCに変換して比較）
                start_dt = pd.to_datetime(start_date, utc=True)
                end_dt = pd.to_datetime(end_date, utc=True)
                self.pivot_purchase_df = self.pivot_purchase_df[
                    (self.pivot_purchase_df['Created time'] >= start_dt) &
                    (self.pivot_purchase_df['Created time'] < end_dt)
                ]

            if not self.pivot_df.empty:
                # 開始年月を保存
                self.pivot_start_year_value = start_year
                self.pivot_start_month_value = start_month

                # 終了年月を計算
                end_year_display = end_year
                end_month_display = end_month - 1
                if end_month_display == 0:
                    end_month_display = 12
                    end_year_display -= 1

                self.result_text.value = f"✅ {start_year}年{start_month}月〜{end_year_display}年{end_month_display}月の{len(self.pivot_df)}件のデータを取得しました"
                self.export_pivot_btn.disabled = False
            else:
                self.result_text.value = "⚠️ 該当期間のデータがありません"
                self.export_pivot_btn.disabled = True

        except Exception as ex:
            self.result_text.value = f"❌ エラー: {str(ex)}"
            self.export_pivot_btn.disabled = True

        self.progress.visible = False
        self.page.update()

    def fetch_daily_data(self, e):
        """日別売上データを取得（選択した日付範囲）"""
        api_key = self.proxy.config.get("notion_api_key", "")
        database_id = self.proxy.config.get("notion_database_id", "")

        if not api_key or not database_id:
            self.show_snackbar("設定タブでNotion API KeyとDatabase IDを設定してください", ft.Colors.RED)
            return

        # 日付の検証
        start_date = self.daily_start_date
        end_date = self.daily_end_date

        if start_date > end_date:
            self.show_snackbar("開始日は終了日より前である必要があります", ft.Colors.RED)
            return

        start_date_str = start_date.strftime("%Y-%m-%d")
        end_date_str = end_date.strftime("%Y-%m-%d")

        self.progress.visible = True
        self.daily_result_text.value = "日別売上データを取得中..."
        self.page.update()

        try:
            notion = Client(auth=api_key)

            # 終了日の翌日を計算（beforeフィルタ用）
            from datetime import timedelta
            end_date_plus_one = end_date + timedelta(days=1)
            end_date_filter = end_date_plus_one.strftime("%Y-%m-%d")

            # 売却済みデータを取得（ページネーション対応）
            all_sold_results = []
            has_more = True
            start_cursor = None

            while has_more:
                query_params = {
                    "database_id": database_id,
                    "filter": {
                        "and": [
                            {
                                "property": "在庫状況",
                                "status": {
                                    "equals": "売却済み"
                                }
                            },
                            {
                                "property": "売却日",
                                "date": {
                                    "on_or_after": start_date_str
                                }
                            },
                            {
                                "property": "売却日",
                                "date": {
                                    "before": end_date_filter
                                }
                            }
                        ]
                    },
                    "page_size": 100
                }

                if start_cursor:
                    query_params["start_cursor"] = start_cursor

                results = notion.databases.query(**query_params)
                all_sold_results.extend(results["results"])
                has_more = results.get("has_more", False)
                start_cursor = results.get("next_cursor")

            # 全結果を辞書形式に変換
            combined_results = {
                "results": all_sold_results
            }

            # データフレームに変換（売却済みデータ）
            self.daily_df = self.parse_notion_results(combined_results)

            # 全データを取得（在庫状況不問、Created time基準）
            all_purchase_results = []
            has_more = True
            start_cursor = None

            while has_more:
                query_params = {
                    "database_id": database_id,
                    "page_size": 100
                }

                if start_cursor:
                    query_params["start_cursor"] = start_cursor

                results = notion.databases.query(**query_params)
                all_purchase_results.extend(results["results"])
                has_more = results.get("has_more", False)
                start_cursor = results.get("next_cursor")

            # 仕入データをDataFrameに変換
            purchase_combined_results = {
                "results": all_purchase_results
            }
            self.daily_purchase_df = self.parse_notion_results(purchase_combined_results)

            # Created timeでフィルタリング（取得後）
            if not self.daily_purchase_df.empty and 'Created time' in self.daily_purchase_df.columns:
                # Created timeを日付型に変換（UTC）
                self.daily_purchase_df['Created time'] = pd.to_datetime(
                    self.daily_purchase_df['Created time'],
                    errors='coerce',
                    utc=True
                )
                # 期間でフィルタ（UTCに変換して比較）
                start_dt = pd.to_datetime(start_date_str, utc=True)
                end_dt = pd.to_datetime(end_date_filter, utc=True)
                self.daily_purchase_df = self.daily_purchase_df[
                    (self.daily_purchase_df['Created time'] >= start_dt) &
                    (self.daily_purchase_df['Created time'] < end_dt)
                ]

            if not self.daily_df.empty:
                # 日付を保存
                self.daily_start_date = start_date_str
                self.daily_end_date = end_date_str

                # 日数を計算
                days_count = (end_date - start_date).days + 1
                if days_count == 1:
                    self.daily_result_text.value = f"✅ {start_date_str}の{len(self.daily_df)}件のデータを取得しました"
                else:
                    self.daily_result_text.value = f"✅ {start_date_str}〜{end_date_str}（{days_count}日間）の{len(self.daily_df)}件のデータを取得しました"

                self.export_daily_btn.disabled = False
            else:
                self.daily_result_text.value = "⚠️ 該当期間のデータがありません"
                self.export_daily_btn.disabled = True

        except Exception as ex:
            self.daily_result_text.value = f"❌ エラー: {str(ex)}"
            self.export_daily_btn.disabled = True

        self.progress.visible = False
        self.page.update()

    def parse_notion_results(self, results):
        """Notion APIの結果をDataFrameに変換"""
        if not results or "results" not in results:
            return pd.DataFrame()

        data = []
        for page in results["results"]:
            properties = page["properties"]
            row = {}

            for prop_name, prop_value in properties.items():
                prop_type = prop_value["type"]

                if prop_type == "title":
                    row[prop_name] = prop_value["title"][0]["plain_text"] if prop_value["title"] else ""
                elif prop_type == "rich_text":
                    row[prop_name] = prop_value["rich_text"][0]["plain_text"] if prop_value["rich_text"] else ""
                elif prop_type == "number":
                    row[prop_name] = prop_value["number"]
                elif prop_type == "select":
                    row[prop_name] = prop_value["select"]["name"] if prop_value["select"] else ""
                elif prop_type == "status":  # status型の処理
                    row[prop_name] = prop_value["status"]["name"] if prop_value["status"] else ""
                elif prop_type == "multi_select":
                    row[prop_name] = ", ".join([item["name"] for item in prop_value["multi_select"]])
                elif prop_type == "date":
                    row[prop_name] = prop_value["date"]["start"] if prop_value["date"] else ""
                elif prop_type == "checkbox":
                    row[prop_name] = prop_value["checkbox"]
                elif prop_type == "url":
                    row[prop_name] = prop_value["url"] or ""
                elif prop_type == "email":
                    row[prop_name] = prop_value["email"] or ""
                elif prop_type == "phone_number":
                    row[prop_name] = prop_value["phone_number"] or ""
                elif prop_type == "formula":
                    # 数式の結果型に応じて処理
                    formula = prop_value.get("formula", {})
                    if formula.get("type") == "string":
                        row[prop_name] = formula.get("string", "")
                    elif formula.get("type") == "number":
                        row[prop_name] = formula.get("number", 0)
                    else:
                        row[prop_name] = str(formula)
                elif prop_type == "rollup":
                    # ロールアップの結果を処理
                    rollup = prop_value.get("rollup", {})
                    rollup_type = rollup.get("type")

                    if rollup_type == "array":
                        array_data = rollup.get("array", [])
                        if array_data:
                            # 配列の各要素を解析
                            values = []
                            for item in array_data:
                                if isinstance(item, dict):
                                    item_type = item.get("type")
                                    # title型（企業名など）
                                    if item_type == "title" and item.get("title"):
                                        values.append(item["title"][0]["plain_text"] if item["title"] else "")
                                    # rich_text型
                                    elif item_type == "rich_text" and item.get("rich_text"):
                                        values.append(item["rich_text"][0]["plain_text"] if item["rich_text"] else "")
                                    # select型
                                    elif item_type == "select" and item.get("select"):
                                        values.append(item["select"]["name"])
                                    # number型
                                    elif item_type == "number":
                                        values.append(str(item.get("number", "")))
                                    # formula型（仕入れ先名、販売媒体名など）
                                    elif item_type == "formula":
                                        formula = item.get("formula", {})
                                        formula_type = formula.get("type")
                                        if formula_type == "string":
                                            string_val = formula.get("string", "")
                                            if string_val:
                                                values.append(string_val)
                                        elif formula_type == "number":
                                            values.append(str(formula.get("number", "")))
                                    else:
                                        values.append(str(item))
                                else:
                                    values.append(str(item))
                            row[prop_name] = ", ".join(filter(None, values))
                        else:
                            row[prop_name] = ""
                    elif rollup_type == "number":
                        row[prop_name] = rollup.get("number", 0)
                    elif rollup_type == "date":
                        date_val = rollup.get("date")
                        row[prop_name] = date_val["start"] if date_val else ""
                    else:
                        row[prop_name] = str(rollup)
                elif prop_type == "relation":
                    # リレーションのIDを取得
                    relations = prop_value.get("relation", [])
                    row[prop_name] = ", ".join([rel["id"] for rel in relations])
                elif prop_type == "people":
                    # ユーザー名を取得
                    people = prop_value.get("people", [])
                    names = []
                    for person in people:
                        # nameフィールドを取得（存在しない場合は空文字）
                        name = person.get("name", "")
                        if name:
                            names.append(name)
                    row[prop_name] = ", ".join(names)
                elif prop_type == "files":
                    # ファイル名を取得
                    files = prop_value.get("files", [])
                    row[prop_name] = ", ".join([file.get("name", "") for file in files])
                elif prop_type == "unique_id":
                    # ユニークIDを取得
                    unique_id = prop_value.get("unique_id", {})
                    prefix = unique_id.get("prefix", "")
                    number = unique_id.get("number", 0)
                    row[prop_name] = f"{prefix}{number}" if prefix else str(number)
                elif prop_type == "created_time":
                    row[prop_name] = prop_value.get("created_time", "")
                else:
                    row[prop_name] = str(prop_value.get(prop_type, ""))

            # ページレベルのCreated timeを追加（仕入日として使用）
            if "created_time" in page:
                row["Created time"] = page["created_time"]

            data.append(row)

        return pd.DataFrame(data)

    def update_data_table(self):
        """データテーブルを更新"""
        if self.df is None or self.df.empty:
            return

        # 列を作成（最大5列まで表示）
        columns = list(self.df.columns)[:5]
        self.data_table.columns = [ft.DataColumn(ft.Text(col)) for col in columns]

        # 行を作成（最大10行まで表示）
        rows = []
        for _, row in self.df.head(10).iterrows():
            cells = [ft.DataCell(ft.Text(str(row.get(col, ""))[:30])) for col in columns]
            rows.append(ft.DataRow(cells=cells))

        self.data_table.rows = rows

    def create_pivot_sections(self, df, start_year, start_month, purchase_df=None):
        """3つのセクションのピボットテーブルを作成（業販+小売を統合）

        Args:
            df: 売却済みデータのDataFrame（売上・利益計算用）
            start_year: 開始年
            start_month: 開始月
            purchase_df: 全データのDataFrame（仕入高計算用、在庫状況不問）
        """
        from datetime import datetime
        import re

        # 月のリストを生成（選択された開始月から12ヶ月分）
        months = []
        current_year = start_year
        current_month = start_month

        for i in range(12):
            months.append(f"{current_year}年{current_month}月")
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

        # データの前処理
        df_clean = df.copy()

        # 売却日から年月を抽出
        def extract_year_month(date_str):
            if pd.isna(date_str) or date_str == "":
                return None
            try:
                # 日付文字列をパース
                date_obj = pd.to_datetime(date_str)
                return f"{date_obj.year}年{date_obj.month}月"
            except:
                return None

        df_clean['売却年月'] = df_clean['売却日'].apply(extract_year_month)

        # 数値フィールドのクリーニング
        def clean_currency(value):
            if pd.isna(value) or value == "":
                return 0
            if isinstance(value, (int, float)):
                return float(value)
            # "￥1,000" のような文字列を数値に変換
            if isinstance(value, str):
                value = value.replace('￥', '').replace(',', '').strip()
                try:
                    return float(value)
                except:
                    return 0
            return 0

        df_clean['売上金_数値'] = df_clean['売上金'].apply(clean_currency)
        df_clean['純利益_数値'] = df_clean['純利益'].apply(clean_currency)
        df_clean['仕入れ原価_数値'] = df_clean['仕入れ原価'].apply(clean_currency)
        df_clean['販売手数料_数値'] = df_clean['販売手数料'].apply(clean_currency)
        df_clean['送料_数値'] = df_clean['送料'].apply(clean_currency)

        # 仕入先と販売媒体のクリーニング
        def clean_company_name(value):
            if pd.isna(value) or value == "":
                return "不明"
            # URL部分を除去
            if isinstance(value, str) and '(https://' in value:
                value = re.sub(r'\s*\(https://.*?\)', '', value)
            return value.strip()

        # カテゴリーのクリーニング（未設定は「小売」）
        def clean_category(value):
            if pd.isna(value) or value == "" or value == "不明":
                return "小売"
            value_str = str(value).strip()
            if value_str in ["市場", "業販", "小売"]:
                return value_str
            return "小売"

        # 仕入れ先名と販売媒体名を使用（ロールアップから取得した企業名）
        df_clean['仕入れ先_clean'] = df_clean['仕入れ先名'].apply(clean_company_name)
        df_clean['販売媒体_clean'] = df_clean['販売媒体名'].apply(clean_company_name)

        # カテゴリー情報を取得（ロールアップから）
        # 販売先カテゴリと仕入れ先カテゴリを使用（最後の「ー」なし）
        if '販売先カテゴリ' in df_clean.columns:
            df_clean['販売先カテゴリー_clean'] = df_clean['販売先カテゴリ'].apply(clean_category)
        else:
            df_clean['販売先カテゴリー_clean'] = "小売"

        if '仕入れ先カテゴリ' in df_clean.columns:
            df_clean['仕入れ先カテゴリー_clean'] = df_clean['仕入れ先カテゴリ'].apply(clean_category)
        else:
            df_clean['仕入れ先カテゴリー_clean'] = "小売"

        # 統合データを作成（仕入先と販売媒体を統合）
        # 仕入先データ
        supplier_data = df_clean.copy()
        supplier_data['企業名'] = supplier_data['仕入れ先_clean']
        supplier_data['カテゴリー'] = supplier_data['仕入れ先カテゴリー_clean']

        # 販売媒体データ
        channel_data = df_clean.copy()
        channel_data['企業名'] = channel_data['販売媒体_clean']
        channel_data['カテゴリー'] = channel_data['販売先カテゴリー_clean']

        # 統合（両方のデータを使用）
        # ※同じレコードが仕入先と販売媒体の両方に計上されるのを避けるため、
        # 仕入先と販売媒体を別々にピボットして後で結合する方式を採用

        # セクション1: 企業別売上（業販+小売）
        # 仕入先別の売上
        pivot_supplier_sales = supplier_data.pivot_table(
            values='売上金_数値',
            index='企業名',
            columns='売却年月',
            aggfunc='sum',
            fill_value=0
        )

        # 販売媒体別の売上
        pivot_channel_sales = channel_data.pivot_table(
            values='売上金_数値',
            index='企業名',
            columns='売却年月',
            aggfunc='sum',
            fill_value=0
        )

        # 統合（重複企業は加算）
        pivot1 = pivot_supplier_sales.add(pivot_channel_sales, fill_value=0)

        # セクション2: 企業別販売利益（業販+小売）
        # 仕入先別の純利益
        pivot_supplier_profit = supplier_data.pivot_table(
            values='純利益_数値',
            index='企業名',
            columns='売却年月',
            aggfunc='sum',
            fill_value=0
        )

        # 販売媒体別の純利益
        pivot_channel_profit = channel_data.pivot_table(
            values='純利益_数値',
            index='企業名',
            columns='売却年月',
            aggfunc='sum',
            fill_value=0
        )

        # 統合（重複企業は加算）
        pivot2 = pivot_supplier_profit.add(pivot_channel_profit, fill_value=0)

        # カテゴリー情報を企業にマッピング
        # 企業名 → カテゴリーのマッピングを作成
        company_category_map = {}

        # 仕入先のカテゴリーマッピング
        for _, row in supplier_data[['企業名', 'カテゴリー']].drop_duplicates().iterrows():
            company = row['企業名']
            category = row['カテゴリー']
            if company and company != '不明':
                # 複数のカテゴリーがある場合は、最初に見つかったものを使用
                if company not in company_category_map:
                    company_category_map[company] = category

        # 販売媒体のカテゴリーマッピング
        for _, row in channel_data[['企業名', 'カテゴリー']].drop_duplicates().iterrows():
            company = row['企業名']
            category = row['カテゴリー']
            if company and company != '不明':
                if company not in company_category_map:
                    company_category_map[company] = category

        # セクション3: 企業別仕入高 - 仕入先別の仕入れ原価
        # 仕入高は在庫状況に関係なく全データから計算
        if purchase_df is not None and not purchase_df.empty:
            # 仕入データの前処理
            purchase_clean = purchase_df.copy()

            # Created timeから年月を抽出（仕入日として扱う）
            def extract_created_month(row):
                # Created timeはparse_notion_resultsで既に処理されている想定
                # もしCreated timeフィールドがない場合は売却日を使う
                if 'Created time' in row and pd.notna(row['Created time']):
                    try:
                        date_obj = pd.to_datetime(row['Created time'])
                        return f"{date_obj.year}年{date_obj.month}月"
                    except:
                        pass
                # フォールバック：売却日を使う
                if '売却日' in row and pd.notna(row['売却日']):
                    try:
                        date_obj = pd.to_datetime(row['売却日'])
                        return f"{date_obj.year}年{date_obj.month}月"
                    except:
                        pass
                return None

            purchase_clean['仕入年月'] = purchase_clean.apply(extract_created_month, axis=1)

            # 仕入れ原価のクリーニング
            purchase_clean['仕入れ原価_数値'] = purchase_clean['仕入れ原価'].apply(clean_currency)

            # 仕入先のクリーニング
            purchase_clean['仕入れ先_clean'] = purchase_clean['仕入れ先'].apply(clean_company_name)

            # 仕入先名を使用（ロールアップから取得した企業名）
            if '仕入れ先名' in purchase_clean.columns:
                purchase_clean['仕入れ先_clean'] = purchase_clean['仕入れ先名'].apply(clean_company_name)

            # カテゴリー情報を取得（仕入データにも追加）
            if '仕入れ先カテゴリ' in purchase_clean.columns:
                purchase_clean['仕入れ先カテゴリー_clean'] = purchase_clean['仕入れ先カテゴリ'].apply(clean_category)
            else:
                purchase_clean['仕入れ先カテゴリー_clean'] = "小売"

            # 仕入データから企業カテゴリーマッピングを更新
            for _, row in purchase_clean[['仕入れ先_clean', '仕入れ先カテゴリー_clean']].drop_duplicates().iterrows():
                company = row['仕入れ先_clean']
                category = row['仕入れ先カテゴリー_clean']
                if company and company != '不明':
                    if company not in company_category_map:
                        company_category_map[company] = category

            # ピボットテーブル作成
            pivot3 = purchase_clean.pivot_table(
                values='仕入れ原価_数値',
                index='仕入れ先_clean',
                columns='仕入年月',
                aggfunc='sum',
                fill_value=0
            )
        else:
            # フォールバック：売却済みデータから計算（後方互換性）
            pivot3 = df_clean.pivot_table(
                values='仕入れ原価_数値',
                index='仕入れ先_clean',
                columns='売却年月',
                aggfunc='sum',
                fill_value=0
            )

        # 月の列を正しい順序で並べ替え、存在しない月は0で埋める
        pivot_list = [pivot1, pivot2, pivot3]
        for i in range(len(pivot_list)):
            pivot = pivot_list[i]
            # すべての月の列を作成（存在しない月は0で埋める）
            for month in months:
                if month not in pivot.columns:
                    pivot[month] = 0
            # 月の順序で並べ替え
            pivot_list[i] = pivot[months]

        # カテゴリー別集計を計算
        # 売上のカテゴリー別集計
        category_sales = {}
        for category in ['市場', '業販', '小売']:
            category_sales[category] = {}
            for month in months:
                category_sales[category][month] = 0
                for company in pivot_list[0].index:
                    if company_category_map.get(company) == category:
                        category_sales[category][month] += pivot_list[0].loc[company, month]

        # 利益のカテゴリー別集計
        category_profit = {}
        for category in ['市場', '業販', '小売']:
            category_profit[category] = {}
            for month in months:
                category_profit[category][month] = 0
                for company in pivot_list[1].index:
                    if company_category_map.get(company) == category:
                        category_profit[category][month] += pivot_list[1].loc[company, month]

        # 仕入れ高のカテゴリー別集計
        category_purchase = {}
        for category in ['市場', '業販', '小売']:
            category_purchase[category] = {}
            for month in months:
                category_purchase[category][month] = 0
                for company in pivot_list[2].index:
                    if company_category_map.get(company) == category:
                        category_purchase[category][month] += pivot_list[2].loc[company, month]

        # 仕入データ（purchase_clean）を返すために保存
        purchase_data_for_mapping = None
        if purchase_df is not None and not purchase_df.empty and 'purchase_clean' in locals():
            purchase_data_for_mapping = purchase_clean

        # 3つのセクションを返す
        return {
            '企業別売上': pivot_list[0],           # 仕入先+販売媒体 × 売上金
            '企業別販売利益': pivot_list[1],       # 仕入先+販売媒体 × 純利益
            '企業別仕入高': pivot_list[2],         # 仕入先 × 仕入れ原価（全データ）
            '企業カテゴリーマッピング': company_category_map,  # 企業 → カテゴリー
            'カテゴリー別売上': category_sales,    # カテゴリー別売上集計
            'カテゴリー別利益': category_profit,   # カテゴリー別利益集計
            'カテゴリー別仕入高': category_purchase,  # カテゴリー別仕入高集計
            '仕入データ': purchase_data_for_mapping  # 仕入データ（担当者マッピング用）
        }

    def create_daily_pivot_sections(self, df, purchase_df=None):
        """日別売上用のピボットテーブルを作成（列は「合計」のみ）

        Args:
            df: 売却済みデータのDataFrame（売上・利益計算用）
            purchase_df: 全データのDataFrame（仕入高計算用、在庫状況不問）
        """
        import re

        # データの前処理
        df_clean = df.copy()

        # 数値フィールドのクリーニング
        def clean_currency(value):
            if pd.isna(value) or value == "":
                return 0
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                value = value.replace('￥', '').replace(',', '').strip()
                try:
                    return float(value)
                except:
                    return 0
            return 0

        df_clean['売上金_数値'] = df_clean['売上金'].apply(clean_currency)
        df_clean['純利益_数値'] = df_clean['純利益'].apply(clean_currency)
        df_clean['仕入れ原価_数値'] = df_clean['仕入れ原価'].apply(clean_currency)

        # 仕入先と販売媒体のクリーニング
        def clean_company_name(value):
            if pd.isna(value) or value == "":
                return "不明"
            if isinstance(value, str) and '(https://' in value:
                value = re.sub(r'\s*\(https://.*?\)', '', value)
            return value.strip()

        # カテゴリーのクリーニング（未設定は「小売」）
        def clean_category(value):
            if pd.isna(value) or value == "" or value == "不明":
                return "小売"
            value_str = str(value).strip()
            if value_str in ["市場", "業販", "小売"]:
                return value_str
            return "小売"

        df_clean['仕入れ先_clean'] = df_clean['仕入れ先名'].apply(clean_company_name) if '仕入れ先名' in df_clean.columns else df_clean['仕入れ先'].apply(clean_company_name)
        df_clean['販売媒体_clean'] = df_clean['販売媒体名'].apply(clean_company_name) if '販売媒体名' in df_clean.columns else "不明"

        # カテゴリー情報を取得
        if '販売先カテゴリ' in df_clean.columns:
            df_clean['販売先カテゴリー_clean'] = df_clean['販売先カテゴリ'].apply(clean_category)
        else:
            df_clean['販売先カテゴリー_clean'] = "小売"

        if '仕入れ先カテゴリ' in df_clean.columns:
            df_clean['仕入れ先カテゴリー_clean'] = df_clean['仕入れ先カテゴリ'].apply(clean_category)
        else:
            df_clean['仕入れ先カテゴリー_clean'] = "小売"

        # 統合データを作成
        supplier_data = df_clean.copy()
        supplier_data['企業名'] = supplier_data['仕入れ先_clean']
        supplier_data['カテゴリー'] = supplier_data['仕入れ先カテゴリー_clean']

        channel_data = df_clean.copy()
        channel_data['企業名'] = channel_data['販売媒体_clean']
        channel_data['カテゴリー'] = channel_data['販売先カテゴリー_clean']

        # 企業別の合計を計算（月列なし、合計のみ）
        # 売上の合計
        supplier_sales = supplier_data.groupby('企業名')['売上金_数値'].sum()
        channel_sales = channel_data.groupby('企業名')['売上金_数値'].sum()
        sales_total = supplier_sales.add(channel_sales, fill_value=0)
        pivot1 = pd.DataFrame({'合計': sales_total})

        # 利益の合計
        supplier_profit = supplier_data.groupby('企業名')['純利益_数値'].sum()
        channel_profit = channel_data.groupby('企業名')['純利益_数値'].sum()
        profit_total = supplier_profit.add(channel_profit, fill_value=0)
        pivot2 = pd.DataFrame({'合計': profit_total})

        # カテゴリーマッピング
        company_category_map = {}
        for _, row in supplier_data[['企業名', 'カテゴリー']].drop_duplicates().iterrows():
            company = row['企業名']
            category = row['カテゴリー']
            if company and company != '不明':
                if company not in company_category_map:
                    company_category_map[company] = category

        for _, row in channel_data[['企業名', 'カテゴリー']].drop_duplicates().iterrows():
            company = row['企業名']
            category = row['カテゴリー']
            if company and company != '不明':
                if company not in company_category_map:
                    company_category_map[company] = category

        # 仕入高の合計
        if purchase_df is not None and not purchase_df.empty:
            purchase_clean = purchase_df.copy()
            purchase_clean['仕入れ原価_数値'] = purchase_clean['仕入れ原価'].apply(clean_currency)
            purchase_clean['仕入れ先_clean'] = purchase_clean['仕入れ先名'].apply(clean_company_name) if '仕入れ先名' in purchase_clean.columns else purchase_clean['仕入れ先'].apply(clean_company_name)

            if '仕入れ先カテゴリ' in purchase_clean.columns:
                purchase_clean['仕入れ先カテゴリー_clean'] = purchase_clean['仕入れ先カテゴリ'].apply(clean_category)
            else:
                purchase_clean['仕入れ先カテゴリー_clean'] = "小売"

            # カテゴリーマッピングを更新
            for _, row in purchase_clean[['仕入れ先_clean', '仕入れ先カテゴリー_clean']].drop_duplicates().iterrows():
                company = row['仕入れ先_clean']
                category = row['仕入れ先カテゴリー_clean']
                if company and company != '不明':
                    if company not in company_category_map:
                        company_category_map[company] = category

            purchase_total = purchase_clean.groupby('仕入れ先_clean')['仕入れ原価_数値'].sum()
            pivot3 = pd.DataFrame({'合計': purchase_total})
        else:
            purchase_total = df_clean.groupby('仕入れ先_clean')['仕入れ原価_数値'].sum()
            pivot3 = pd.DataFrame({'合計': purchase_total})

        # カテゴリー別集計
        category_sales = {}
        category_profit = {}
        category_purchase = {}

        for category in ['市場', '業販', '小売']:
            category_sales[category] = {'合計': 0}
            category_profit[category] = {'合計': 0}
            category_purchase[category] = {'合計': 0}

            for company in pivot1.index:
                if company_category_map.get(company) == category:
                    category_sales[category]['合計'] += pivot1.loc[company, '合計']

            for company in pivot2.index:
                if company_category_map.get(company) == category:
                    category_profit[category]['合計'] += pivot2.loc[company, '合計']

            for company in pivot3.index:
                if company_category_map.get(company) == category:
                    category_purchase[category]['合計'] += pivot3.loc[company, '合計']

        # 仕入データを保存
        purchase_data_for_mapping = None
        if purchase_df is not None and not purchase_df.empty:
            purchase_data_for_mapping = purchase_clean

        return {
            '企業別売上': pivot1,
            '企業別販売利益': pivot2,
            '企業別仕入高': pivot3,
            '企業カテゴリーマッピング': company_category_map,
            'カテゴリー別売上': category_sales,
            'カテゴリー別利益': category_profit,
            'カテゴリー別仕入高': category_purchase,
            '仕入データ': purchase_data_for_mapping
        }

    def export_pivot_excel(self, e):
        """ピボット形式の財務集計Excelを保存"""
        if self.pivot_df is None or self.pivot_df.empty:
            return

        # 終了年月を計算
        start_year = self.pivot_start_year_value
        start_month = self.pivot_start_month_value
        end_month = start_month + 11
        end_year = start_year
        if end_month > 12:
            end_year = start_year + 1
            end_month = end_month - 12

        file_name = f"財務集計_{start_year}年{start_month}月-{end_year}年{end_month}月.xlsx"

        def save_file(e: ft.FilePickerResultEvent):
            if e.path:
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

                    # ピボットセクションを作成
                    sections = self.create_pivot_sections(
                        self.pivot_df,
                        self.pivot_start_year_value,
                        self.pivot_start_month_value,
                        self.pivot_purchase_df if hasattr(self, 'pivot_purchase_df') else None
                    )

                    # 全体合算データを作成
                    summary_data = None
                    if not self.pivot_df.empty:
                        import re
                        from datetime import datetime

                        # 月リストを生成
                        months = []
                        current_year = self.pivot_start_year_value
                        current_month = self.pivot_start_month_value
                        for i in range(12):
                            months.append(f"{current_year}年{current_month}月")
                            current_month += 1
                            if current_month > 12:
                                current_month = 1
                                current_year += 1

                        # データクリーニング関数
                        def clean_currency(value):
                            if pd.isna(value) or value == "":
                                return 0
                            if isinstance(value, (int, float)):
                                return float(value)
                            if isinstance(value, str):
                                value = value.replace('￥', '').replace(',', '').strip()
                                try:
                                    return float(value)
                                except:
                                    return 0
                            return 0

                        # 売却日から年月を抽出
                        def extract_year_month(date_str):
                            if pd.isna(date_str) or date_str == "":
                                return None
                            try:
                                date_obj = pd.to_datetime(date_str)
                                return f"{date_obj.year}年{date_obj.month}月"
                            except:
                                return None

                        df_temp = self.pivot_df.copy()
                        df_temp['売却年月'] = df_temp['売却日'].apply(extract_year_month)
                        df_temp['売上金_数値'] = df_temp['売上金'].apply(clean_currency)
                        df_temp['仕入れ原価_数値'] = df_temp['仕入れ原価'].apply(clean_currency)
                        df_temp['販売手数料_数値'] = df_temp['販売手数料'].apply(clean_currency)
                        df_temp['送料_数値'] = df_temp['送料'].apply(clean_currency)
                        df_temp['純利益_数値'] = df_temp['純利益'].apply(clean_currency)

                        summary_rows = {}
                        for month in months:
                            month_data = df_temp[df_temp['売却年月'] == month]
                            summary_rows.setdefault('売上', {})[month] = month_data['売上金_数値'].sum()
                            summary_rows.setdefault('原価', {})[month] = month_data['仕入れ原価_数値'].sum()
                            summary_rows.setdefault('販売手数料', {})[month] = month_data['販売手数料_数値'].sum()
                            summary_rows.setdefault('送料', {})[month] = month_data['送料_数値'].sum()
                            summary_rows.setdefault('販売利益', {})[month] = month_data['純利益_数値'].sum()
                            summary_rows.setdefault('粗利', {})[month] = summary_rows['売上'][month] - summary_rows['原価'][month]

                        summary_data = pd.DataFrame(summary_rows).T
                        summary_data = summary_data.reindex(['売上', '原価', '粗利', '販売手数料', '送料', '販売利益'])
                        summary_data = summary_data[months]

                    # 担当者→企業リストのマッピングを作成
                    assignee_company_mapping = {}

                    # 売却済みデータから作業担当を取得
                    if '作業担当' in self.pivot_df.columns:
                        import re

                        # 仕入先の担当者グループ化（仕入れ先名を使用）
                        if '仕入れ先名' in self.pivot_df.columns and '作業担当' in self.pivot_df.columns:
                            for _, row in self.pivot_df[['仕入れ先名', '作業担当']].drop_duplicates().iterrows():
                                supplier = row['仕入れ先名']
                                assignee = row['作業担当']
                                if pd.notna(supplier) and supplier and supplier != '不明' and pd.notna(assignee):
                                    # URL除去してクリーン名を取得
                                    clean_name = supplier
                                    if isinstance(supplier, str) and '(https://' in supplier:
                                        clean_name = re.sub(r'\s*\(https://.*?\)', '', supplier).strip()

                                    if assignee not in assignee_company_mapping:
                                        assignee_company_mapping[assignee] = []
                                    if clean_name not in assignee_company_mapping[assignee]:
                                        assignee_company_mapping[assignee].append(clean_name)

                        # 販売媒体の担当者グループ化（販売媒体名を使用）
                        if '販売媒体名' in self.pivot_df.columns and '作業担当' in self.pivot_df.columns:
                            for _, row in self.pivot_df[['販売媒体名', '作業担当']].drop_duplicates().iterrows():
                                channel = row['販売媒体名']
                                assignee = row['作業担当']
                                if pd.notna(channel) and channel and channel != '不明' and pd.notna(assignee):
                                    # URL除去してクリーン名を取得
                                    clean_name = channel
                                    if isinstance(channel, str) and '(https://' in channel:
                                        clean_name = re.sub(r'\s*\(https://.*?\)', '', channel).strip()

                                    if assignee not in assignee_company_mapping:
                                        assignee_company_mapping[assignee] = []
                                    if clean_name not in assignee_company_mapping[assignee]:
                                        assignee_company_mapping[assignee].append(clean_name)

                    # 仕入データからも担当者情報を取得（全在庫状況のデータ）
                    purchase_data = sections.get('仕入データ')
                    if purchase_data is not None and not purchase_data.empty:
                        import re
                        if '作業担当' in purchase_data.columns and '仕入れ先_clean' in purchase_data.columns:
                            for _, row in purchase_data[['仕入れ先_clean', '作業担当']].drop_duplicates().iterrows():
                                supplier = row['仕入れ先_clean']
                                assignee = row['作業担当']
                                if pd.notna(supplier) and supplier and supplier != '不明' and pd.notna(assignee):
                                    if assignee not in assignee_company_mapping:
                                        assignee_company_mapping[assignee] = []
                                    if supplier not in assignee_company_mapping[assignee]:
                                        assignee_company_mapping[assignee].append(supplier)

                    # Excelワークブックを作成
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "財務集計"

                    current_row = 1

                    # 色の定義
                    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 灰色
                    total_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # 薄橙色

                    # 担当者ごとの色パターン（6色）
                    assignee_colors = [
                        PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # 薄青
                        PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid"),  # 薄ピンク
                        PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),  # 薄緑
                        PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),  # 薄オレンジ
                        PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),  # 薄紫
                        PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),  # 薄黄色
                    ]

                    # 全体合算セクションを最初に書き込み
                    if summary_data is not None and not summary_data.empty:
                        # セクションタイトル
                        ws.cell(row=current_row, column=1, value="全体合算")
                        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                        current_row += 1

                        # ヘッダー行（月の列） - 灰色背景
                        ws.cell(row=current_row, column=1, value="内訳")
                        ws.cell(row=current_row, column=1).fill = header_fill
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        for col_idx, month in enumerate(summary_data.columns, start=2):
                            ws.cell(row=current_row, column=col_idx, value=month)
                            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
                            ws.cell(row=current_row, column=col_idx).fill = header_fill

                        # 「計」列 - 灰色背景
                        ws.cell(row=current_row, column=len(summary_data.columns) + 2, value="計")
                        ws.cell(row=current_row, column=len(summary_data.columns) + 2).font = Font(bold=True)
                        ws.cell(row=current_row, column=len(summary_data.columns) + 2).alignment = Alignment(horizontal='center')
                        ws.cell(row=current_row, column=len(summary_data.columns) + 2).fill = header_fill
                        current_row += 1

                        # データ行
                        for item_name in summary_data.index:
                            ws.cell(row=current_row, column=1, value=item_name)
                            ws.cell(row=current_row, column=1).fill = total_fill
                            ws.cell(row=current_row, column=1).font = Font(bold=True)
                            row_total = 0
                            for col_idx, month in enumerate(summary_data.columns, start=2):
                                value = summary_data.loc[item_name, month]
                                ws.cell(row=current_row, column=col_idx, value=value)
                                ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                ws.cell(row=current_row, column=col_idx).fill = total_fill
                                ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                                row_total += value

                            # 計列
                            ws.cell(row=current_row, column=len(summary_data.columns) + 2, value=row_total)
                            ws.cell(row=current_row, column=len(summary_data.columns) + 2).number_format = '#,##0'
                            ws.cell(row=current_row, column=len(summary_data.columns) + 2).fill = total_fill
                            ws.cell(row=current_row, column=len(summary_data.columns) + 2).font = Font(bold=True)
                            current_row += 1

                        current_row += 1  # 全体合算と各セクションの間に空行

                    # 各セクションを順番に書き込み
                    # カテゴリー関連のデータを取得
                    company_category_map = sections.get('企業カテゴリーマッピング', {})
                    category_sales = sections.get('カテゴリー別売上', {})
                    category_profit = sections.get('カテゴリー別利益', {})
                    category_purchase = sections.get('カテゴリー別仕入高', {})

                    # カテゴリー別の背景色
                    category_fill = PatternFill(start_color="E0F2F7", end_color="E0F2F7", fill_type="solid")  # 薄水色

                    # 実際のピボットセクションのみを処理（メタデータを除外）
                    pivot_sections = {k: v for k, v in sections.items()
                                     if k in ['企業別売上', '企業別販売利益', '企業別仕入高']}

                    for section_name, pivot_df in pivot_sections.items():
                        # セクションタイトル
                        ws.cell(row=current_row, column=1, value=section_name)
                        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                        current_row += 1

                        # ヘッダー行（月の列） - 灰色背景
                        ws.cell(row=current_row, column=1, value="")  # 左上は空欄
                        ws.cell(row=current_row, column=1).fill = header_fill
                        for col_idx, month in enumerate(pivot_df.columns, start=2):
                            ws.cell(row=current_row, column=col_idx, value=month)
                            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                            ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='center')
                            ws.cell(row=current_row, column=col_idx).fill = header_fill

                        # 「計」列 - 灰色背景
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value="計")
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).alignment = Alignment(horizontal='center')
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = header_fill
                        current_row += 1

                        # 担当者ごとにグループ化して表示
                        if assignee_company_mapping:
                            assignee_idx = 0
                            for assignee, companies in sorted(assignee_company_mapping.items()):
                                # 担当者の色を選択（循環）
                                assignee_fill = assignee_colors[assignee_idx % len(assignee_colors)]
                                assignee_idx += 1

                                # その担当者の企業データを表示 + 小計の計算
                                assignee_subtotal_by_month = {month: 0 for month in pivot_df.columns}
                                assignee_grand_total = 0

                                # 全セクションで企業名を表示するように変更（入庫も含む）
                                for company in companies:
                                    if company in pivot_df.index:
                                        # 企業別仕入高セクションでも企業名を表示する
                                        ws.cell(row=current_row, column=1, value=company)
                                        row_total = 0
                                        for col_idx, month in enumerate(pivot_df.columns, start=2):
                                            value = pivot_df.loc[company, month]
                                            ws.cell(row=current_row, column=col_idx, value=value)
                                            ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                            row_total += value
                                            assignee_subtotal_by_month[month] += value

                                        # 計列
                                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=row_total)
                                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                                        assignee_grand_total += row_total
                                        current_row += 1

                                # 担当者の小計行（担当者名+小計）
                                label_suffix = "粗利計" if section_name == '企業別販売利益' else "計"
                                ws.cell(row=current_row, column=1, value=f"{assignee}{label_suffix}")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = assignee_fill
                                for col_idx, month in enumerate(pivot_df.columns, start=2):
                                    subtotal_value = assignee_subtotal_by_month[month]
                                    ws.cell(row=current_row, column=col_idx, value=subtotal_value)
                                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                    ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                                    ws.cell(row=current_row, column=col_idx).fill = assignee_fill

                                # 小計の計列
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=assignee_grand_total)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = assignee_fill
                                current_row += 1
                        else:
                            # 担当者マッピングがない場合は従来通り企業を表示
                            for company in pivot_df.index:
                                ws.cell(row=current_row, column=1, value=company)
                                row_total = 0
                                for col_idx, month in enumerate(pivot_df.columns, start=2):
                                    value = pivot_df.loc[company, month]
                                    ws.cell(row=current_row, column=col_idx, value=value)
                                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                    row_total += value

                                # 計列
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=row_total)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                                current_row += 1

                        # カテゴリー別合計行を追加（売上と利益のセクションのみ）
                        if section_name == '企業別売上' and category_sales:
                            for category in ['市場', '業販', '小売']:
                                ws.cell(row=current_row, column=1, value=f"{category}合計")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = category_fill

                                category_row_total = 0
                                for col_idx, month in enumerate(pivot_df.columns, start=2):
                                    category_value = category_sales.get(category, {}).get(month, 0)
                                    ws.cell(row=current_row, column=col_idx, value=category_value)
                                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                    ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                                    ws.cell(row=current_row, column=col_idx).fill = category_fill
                                    category_row_total += category_value

                                # カテゴリー合計の計列
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=category_row_total)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = category_fill
                                current_row += 1

                        elif section_name == '企業別販売利益' and category_profit:
                            for category in ['市場', '業販', '小売']:
                                ws.cell(row=current_row, column=1, value=f"{category}合計")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = category_fill

                                category_row_total = 0
                                for col_idx, month in enumerate(pivot_df.columns, start=2):
                                    category_value = category_profit.get(category, {}).get(month, 0)
                                    ws.cell(row=current_row, column=col_idx, value=category_value)
                                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                    ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                                    ws.cell(row=current_row, column=col_idx).fill = category_fill
                                    category_row_total += category_value

                                # カテゴリー合計の計列
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=category_row_total)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = category_fill
                                current_row += 1

                        elif section_name == '企業別仕入高' and category_purchase:
                            for category in ['市場', '業販', '小売']:
                                ws.cell(row=current_row, column=1, value=f"{category}合計")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = category_fill

                                category_row_total = 0
                                for col_idx, month in enumerate(pivot_df.columns, start=2):
                                    category_value = category_purchase.get(category, {}).get(month, 0)
                                    ws.cell(row=current_row, column=col_idx, value=category_value)
                                    ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                                    ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                                    ws.cell(row=current_row, column=col_idx).fill = category_fill
                                    category_row_total += category_value

                                # カテゴリー合計の計列
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=category_row_total)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                                ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = category_fill
                                current_row += 1

                        # 合計行（薄橙色背景）
                        ws.cell(row=current_row, column=1, value="合計")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        ws.cell(row=current_row, column=1).fill = total_fill
                        for col_idx, month in enumerate(pivot_df.columns, start=2):
                            col_total = pivot_df[month].sum()
                            ws.cell(row=current_row, column=col_idx, value=col_total)
                            ws.cell(row=current_row, column=col_idx).number_format = '#,##0'
                            ws.cell(row=current_row, column=col_idx).font = Font(bold=True)
                            ws.cell(row=current_row, column=col_idx).fill = total_fill

                        # 合計の計列
                        grand_total = pivot_df.values.sum()
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2, value=grand_total)
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).number_format = '#,##0'
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).font = Font(bold=True)
                        ws.cell(row=current_row, column=len(pivot_df.columns) + 2).fill = total_fill
                        current_row += 1

                        current_row += 1  # セクション間に空行

                    # 列幅調整
                    ws.column_dimensions['A'].width = 20
                    # 最初のセクションの列数を取得
                    num_cols = len(list(pivot_sections.values())[0].columns) + 2
                    for col_idx in range(2, num_cols + 1):
                        # A=1, B=2, ... Z=26, AA=27, AB=28, ...
                        if col_idx <= 26:
                            col_letter = chr(64 + col_idx)
                        else:
                            first_letter = chr(64 + (col_idx - 1) // 26)
                            second_letter = chr(65 + (col_idx - 1) % 26)
                            col_letter = first_letter + second_letter
                        ws.column_dimensions[col_letter].width = 12

                    # 保存
                    wb.save(e.path)
                    self.show_snackbar(f"保存しました: {e.path}", ft.Colors.GREEN)

                except Exception as ex:
                    import traceback
                    error_detail = traceback.format_exc()
                    self.show_snackbar(f"保存エラー: {str(ex)}", ft.Colors.RED)
                    print(error_detail)  # デバッグ用

        file_picker = ft.FilePicker(on_result=save_file)
        self.page.overlay.append(file_picker)
        self.page.update()

        file_picker.save_file(
            dialog_title="財務集計Excelを保存",
            file_name=file_name,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx"]
        )

    def show_snackbar(self, message: str, color):
        """スナックバーを表示"""
        snack = ft.SnackBar(
            content=ft.Text(message),
            bgcolor=color
        )
        self.page.snack_bar = snack
        self.page.snack_bar.open = True
        self.page.update()

    def export_daily_excel(self, e):
        """日別売上Excelを保存（export_pivot_excelを流用）"""
        if self.daily_df is None or self.daily_df.empty:
            return

        # ファイル名を生成
        if self.daily_start_date == self.daily_end_date:
            file_name = f"日別売上_{self.daily_start_date}.xlsx"
        else:
            file_name = f"日別売上_{self.daily_start_date}_{self.daily_end_date}.xlsx"

        def save_file(e: ft.FilePickerResultEvent):
            if e.path:
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

                    # ピボットセクションを作成（日別用）
                    sections = self.create_daily_pivot_sections(
                        self.daily_df,
                        self.daily_purchase_df if hasattr(self, 'daily_purchase_df') else None
                    )

                    # Excelワークブックを作成
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "日別売上"

                    current_row = 1

                    # 色の定義
                    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    total_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    category_fill = PatternFill(start_color="E0F2F7", end_color="E0F2F7", fill_type="solid")

                    assignee_colors = [
                        PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
                        PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid"),
                        PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
                        PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
                        PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),
                        PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),
                    ]

                    # タイトル行
                    if self.daily_start_date == self.daily_end_date:
                        ws.cell(row=current_row, column=1, value=f"日別売上集計（{self.daily_start_date}）")
                    else:
                        ws.cell(row=current_row, column=1, value=f"日別売上集計（{self.daily_start_date}〜{self.daily_end_date}）")
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16)
                    current_row += 2

                    # カテゴリー関連のデータを取得
                    company_category_map = sections.get('企業カテゴリーマッピング', {})
                    category_sales = sections.get('カテゴリー別売上', {})
                    category_profit = sections.get('カテゴリー別利益', {})
                    category_purchase = sections.get('カテゴリー別仕入高', {})

                    # 担当者マッピングを作成
                    assignee_company_mapping = {}
                    import re

                    if '作業担当' in self.daily_df.columns:
                        if '仕入れ先名' in self.daily_df.columns:
                            for _, row in self.daily_df[['仕入れ先名', '作業担当']].drop_duplicates().iterrows():
                                supplier = row['仕入れ先名']
                                assignee = row['作業担当']
                                if pd.notna(supplier) and supplier and supplier != '不明' and pd.notna(assignee):
                                    clean_name = supplier
                                    if isinstance(supplier, str) and '(https://' in supplier:
                                        clean_name = re.sub(r'\s*\(https://.*?\)', '', supplier).strip()
                                    if assignee not in assignee_company_mapping:
                                        assignee_company_mapping[assignee] = []
                                    if clean_name not in assignee_company_mapping[assignee]:
                                        assignee_company_mapping[assignee].append(clean_name)

                        if '販売媒体名' in self.daily_df.columns:
                            for _, row in self.daily_df[['販売媒体名', '作業担当']].drop_duplicates().iterrows():
                                channel = row['販売媒体名']
                                assignee = row['作業担当']
                                if pd.notna(channel) and channel and channel != '不明' and pd.notna(assignee):
                                    clean_name = channel
                                    if isinstance(channel, str) and '(https://' in channel:
                                        clean_name = re.sub(r'\s*\(https://.*?\)', '', channel).strip()
                                    if assignee not in assignee_company_mapping:
                                        assignee_company_mapping[assignee] = []
                                    if clean_name not in assignee_company_mapping[assignee]:
                                        assignee_company_mapping[assignee].append(clean_name)

                    # 仕入データからも担当者情報を取得
                    purchase_data = sections.get('仕入データ')
                    if purchase_data is not None and not purchase_data.empty:
                        if '作業担当' in purchase_data.columns and '仕入れ先_clean' in purchase_data.columns:
                            for _, row in purchase_data[['仕入れ先_clean', '作業担当']].drop_duplicates().iterrows():
                                supplier = row['仕入れ先_clean']
                                assignee = row['作業担当']
                                if pd.notna(supplier) and supplier and supplier != '不明' and pd.notna(assignee):
                                    if assignee not in assignee_company_mapping:
                                        assignee_company_mapping[assignee] = []
                                    if supplier not in assignee_company_mapping[assignee]:
                                        assignee_company_mapping[assignee].append(supplier)

                    # 各セクションを書き込み
                    pivot_sections = {k: v for k, v in sections.items()
                                     if k in ['企業別売上', '企業別販売利益', '企業別仕入高']}

                    for section_name, pivot_df in pivot_sections.items():
                        # セクションタイトル
                        ws.cell(row=current_row, column=1, value=section_name)
                        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                        current_row += 1

                        # ヘッダー行
                        ws.cell(row=current_row, column=1, value="")
                        ws.cell(row=current_row, column=1).fill = header_fill
                        ws.cell(row=current_row, column=2, value="合計")
                        ws.cell(row=current_row, column=2).font = Font(bold=True)
                        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
                        ws.cell(row=current_row, column=2).fill = header_fill
                        current_row += 1

                        # 担当者ごとにグループ化して表示
                        if assignee_company_mapping:
                            assignee_idx = 0
                            for assignee, companies in sorted(assignee_company_mapping.items()):
                                assignee_fill = assignee_colors[assignee_idx % len(assignee_colors)]
                                assignee_idx += 1

                                assignee_total = 0

                                # 全セクションで企業名を表示（入庫も含む）
                                for company in companies:
                                    if company in pivot_df.index:
                                        ws.cell(row=current_row, column=1, value=company)
                                        value = pivot_df.loc[company, '合計']
                                        ws.cell(row=current_row, column=2, value=value)
                                        ws.cell(row=current_row, column=2).number_format = '#,##0'
                                        assignee_total += value
                                        current_row += 1

                                # 担当者の小計行
                                label_suffix = "粗利計" if section_name == '企業別販売利益' else "計"
                                ws.cell(row=current_row, column=1, value=f"{assignee}{label_suffix}")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = assignee_fill
                                ws.cell(row=current_row, column=2, value=assignee_total)
                                ws.cell(row=current_row, column=2).number_format = '#,##0'
                                ws.cell(row=current_row, column=2).font = Font(bold=True)
                                ws.cell(row=current_row, column=2).fill = assignee_fill
                                current_row += 1
                        else:
                            for company in pivot_df.index:
                                ws.cell(row=current_row, column=1, value=company)
                                value = pivot_df.loc[company, '合計']
                                ws.cell(row=current_row, column=2, value=value)
                                ws.cell(row=current_row, column=2).number_format = '#,##0'
                                current_row += 1

                        # カテゴリー別合計行を追加
                        if section_name == '企業別売上' and category_sales:
                            for category in ['市場', '業販', '小売']:
                                ws.cell(row=current_row, column=1, value=f"{category}合計")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = category_fill
                                category_value = category_sales.get(category, {}).get('合計', 0)
                                ws.cell(row=current_row, column=2, value=category_value)
                                ws.cell(row=current_row, column=2).number_format = '#,##0'
                                ws.cell(row=current_row, column=2).font = Font(bold=True)
                                ws.cell(row=current_row, column=2).fill = category_fill
                                current_row += 1

                        elif section_name == '企業別販売利益' and category_profit:
                            for category in ['市場', '業販', '小売']:
                                ws.cell(row=current_row, column=1, value=f"{category}合計")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = category_fill
                                category_value = category_profit.get(category, {}).get('合計', 0)
                                ws.cell(row=current_row, column=2, value=category_value)
                                ws.cell(row=current_row, column=2).number_format = '#,##0'
                                ws.cell(row=current_row, column=2).font = Font(bold=True)
                                ws.cell(row=current_row, column=2).fill = category_fill
                                current_row += 1

                        elif section_name == '企業別仕入高' and category_purchase:
                            for category in ['市場', '業販', '小売']:
                                ws.cell(row=current_row, column=1, value=f"{category}合計")
                                ws.cell(row=current_row, column=1).font = Font(bold=True)
                                ws.cell(row=current_row, column=1).fill = category_fill
                                category_value = category_purchase.get(category, {}).get('合計', 0)
                                ws.cell(row=current_row, column=2, value=category_value)
                                ws.cell(row=current_row, column=2).number_format = '#,##0'
                                ws.cell(row=current_row, column=2).font = Font(bold=True)
                                ws.cell(row=current_row, column=2).fill = category_fill
                                current_row += 1

                        # 合計行
                        ws.cell(row=current_row, column=1, value="合計")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        ws.cell(row=current_row, column=1).fill = total_fill
                        grand_total = pivot_df['合計'].sum()
                        ws.cell(row=current_row, column=2, value=grand_total)
                        ws.cell(row=current_row, column=2).number_format = '#,##0'
                        ws.cell(row=current_row, column=2).font = Font(bold=True)
                        ws.cell(row=current_row, column=2).fill = total_fill
                        current_row += 1

                        current_row += 1  # セクション間に空行

                    # 列幅調整
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 15

                    # 保存
                    wb.save(e.path)
                    self.show_snackbar(f"保存しました: {e.path}", ft.Colors.GREEN)

                except Exception as ex:
                    import traceback
                    error_detail = traceback.format_exc()
                    self.show_snackbar(f"保存エラー: {str(ex)}", ft.Colors.RED)
                    print(error_detail)

        file_picker = ft.FilePicker(on_result=save_file)
        self.page.overlay.append(file_picker)
        self.page.update()

        file_picker.save_file(
            dialog_title="日別売上Excelを保存",
            file_name=file_name,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx"]
        )

    def create_tab(self) -> ft.Tab:
        """タブを作成"""
        # DatePickerをページのoverlayに追加
        self.page.overlay.extend([
            self.pivot_date_picker,
            self.daily_start_date_picker_dialog,
            self.daily_end_date_picker_dialog,
        ])

        return ft.Tab(
            text="Excel出力",
            icon=ft.Icons.TABLE_CHART,
            content=ft.Container(
                padding=ft.padding.all(20),
                content=ft.Column([
                    # 期間選択カード
                    ft.Card(
                        content=ft.Container(
                            padding=ft.padding.all(20),
                            content=ft.Column([
                                ft.Text("財務集計期間選択", size=18, weight=ft.FontWeight.BOLD),
                                ft.Divider(),
                                ft.Text("開始年月（12ヶ月分を集計）", size=14, weight=ft.FontWeight.BOLD),
                                self.pivot_date_button,
                                ft.Text("※「在庫状況」が「売却済み」のデータのみ取得します",
                                       size=12, color=ft.Colors.GREY_600),
                            ])
                        )
                    ),
                    # 操作ボタン
                    ft.Row([
                        self.fetch_pivot_btn,
                        self.export_pivot_btn,
                    ]),
                    self.progress,
                    self.result_text,
                    # 日別売上カード
                    ft.Divider(height=30, color=ft.Colors.TRANSPARENT),
                    ft.Card(
                        content=ft.Container(
                            padding=ft.padding.all(20),
                            content=ft.Column([
                                ft.Text("日別売上集計", size=18, weight=ft.FontWeight.BOLD),
                                ft.Divider(),
                                ft.Text("日付範囲を選択", size=14, weight=ft.FontWeight.BOLD),
                                ft.Row([
                                    self.daily_start_date_button,
                                    self.daily_end_date_button,
                                ]),
                                ft.Text("※デフォルトは今日の日付（1日分）です",
                                       size=12, color=ft.Colors.GREY_600),
                            ])
                        )
                    ),
                    # 日別売上操作ボタン
                    ft.Row([
                        self.fetch_daily_btn,
                        self.export_daily_btn,
                    ]),
                    self.daily_result_text,
                ], scroll=ft.ScrollMode.AUTO)
            )
        )


def create_export_tab(proxy, page: ft.Page) -> tuple[ft.Tab, ExportTab]:
    """Excel出力タブを作成して返す"""
    export_tab = ExportTab(proxy, page)
    return export_tab.create_tab(), export_tab
