import pandas as pd
import time
from datetime import datetime, timedelta
from urllib.parse import urlencode
from notion_client import Client
from notion_client.errors import APIResponseError
from typing import List, Dict, Any, Optional, Tuple, Callable
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from .schemas import SoldRecord, PurchaseRecord, DailySoldRecord, DailyPurchaseRecord


class FetchCancelled(Exception):
    """データ取得がキャンセルされた場合の例外"""
    pass


class ExportService:
    # レート制限の設定
    MAX_RETRIES = 5
    INITIAL_RETRY_DELAY = 1.0  # 秒
    MAX_RETRY_DELAY = 60.0  # 秒
    REQUEST_DELAY = 0.35  # リクエスト間の待機時間（秒）- Notion APIは約3リクエスト/秒

    def __init__(self, api_key: str, database_id: str):
        self.api_key = api_key
        self.database_id = database_id
        if api_key:
            self.notion = Client(auth=api_key, notion_version="2022-06-28")
        else:
            self.notion = None

        # キャンセルフラグ
        self._cancelled = False
        # 進捗コールバック
        self._progress_callback: Optional[Callable[[int, int, str], None]] = None
        # プロパティ名 → ID のキャッシュ（filter_properties 用）
        self._property_id_map: Optional[Dict[str, str]] = None

    def set_progress_callback(self, callback: Optional[Callable[[int, int, str], None]]):
        """進捗コールバックを設定

        Args:
            callback: (current, total, message) を受け取るコールバック関数
        """
        self._progress_callback = callback

    def cancel(self):
        """データ取得をキャンセル"""
        self._cancelled = True

    def reset_cancel(self):
        """キャンセルフラグをリセット"""
        self._cancelled = False

    def _check_cancelled(self):
        """キャンセルされていたら例外を発生"""
        if self._cancelled:
            raise FetchCancelled("データ取得がキャンセルされました")

    def _report_progress(self, current: int, total: int, message: str):
        """進捗を報告"""
        if self._progress_callback:
            self._progress_callback(current, total, message)

    def _query_with_retry(
        self,
        query_params: Dict[str, Any],
        filter_properties: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """リトライロジック付きでNotion APIをクエリ

        Args:
            query_params: databases.queryに渡すパラメータ
            filter_properties: 取得するプロパティのID一覧。指定時は
                ?filter_properties=<id>&... をクエリ文字列で渡し、
                レスポンスに含めるプロパティを絞り込む（転送量の削減）。

        Returns:
            APIレスポンス

        Raises:
            FetchCancelled: キャンセルされた場合
            APIResponseError: リトライ回数を超えた場合
        """
        retry_delay = self.INITIAL_RETRY_DELAY
        last_error = None

        for attempt in range(self.MAX_RETRIES):
            self._check_cancelled()

            try:
                params = dict(query_params)
                database_id = params.pop("database_id")
                path = f"databases/{database_id}/query"
                if filter_properties:
                    query_string = urlencode(
                        [("filter_properties", pid) for pid in filter_properties]
                    )
                    path = f"{path}?{query_string}"
                result = self.notion.request(
                    path=path,
                    method="POST",
                    body=params,
                )
                # 成功したらリクエスト間の待機
                time.sleep(self.REQUEST_DELAY)
                return result

            except APIResponseError as e:
                if e.status == 429:  # Rate limited
                    last_error = e
                    # Retry-Afterヘッダーがあれば使用
                    retry_after = getattr(e, 'retry_after', None)
                    wait_time = float(retry_after) if retry_after else retry_delay
                    wait_time = min(wait_time, self.MAX_RETRY_DELAY)

                    self._report_progress(
                        -1, -1,
                        f"レート制限のため {wait_time:.1f}秒待機中... (リトライ {attempt + 1}/{self.MAX_RETRIES})"
                    )

                    # 待機中もキャンセルチェック
                    wait_end = time.time() + wait_time
                    while time.time() < wait_end:
                        self._check_cancelled()
                        time.sleep(0.5)

                    # エクスポネンシャルバックオフ
                    retry_delay = min(retry_delay * 2, self.MAX_RETRY_DELAY)
                else:
                    # 429以外のエラーは即座に再スロー
                    raise

        # リトライ回数を超過
        if last_error:
            raise last_error
        raise RuntimeError(f"リトライ回数({self.MAX_RETRIES})を超過しました")

    def _get_property_ids(self, names: List[str]) -> List[str]:
        """プロパティ名のリストを対応する property ID のリストに解決する。

        DB スキーマ（databases/{id}）を 1 回だけ取得してキャッシュし、
        プロパティ名 → id のマップを作る。filter_properties 用。

        Args:
            names: 解決したいプロパティ名のリスト

        Returns:
            解決できた property ID のリスト。1件も解決できない、または
            スキーマ取得に失敗した場合は空リストを返す（呼び出し側は
            絞り込みなしにフォールバックする）。
        """
        if self._property_id_map is None:
            try:
                db = self.notion.request(
                    path=f"databases/{self.database_id}",
                    method="GET",
                )
                self._property_id_map = {
                    name.strip(): prop.get("id")
                    for name, prop in db.get("properties", {}).items()
                }
            except Exception as e:
                print(f"プロパティID取得に失敗しました（絞り込みなしで続行）: {e}")
                self._property_id_map = {}

        ids = [
            self._property_id_map[name]
            for name in names
            if name in self._property_id_map and self._property_id_map[name]
        ]
        return ids

    def fetch_sales_data(self, start_date: str, end_date: str) -> List[SoldRecord]:
        """Fetch sold items from Notion within the date range

        Args:
            start_date: 開始日 (YYYY-MM-DD形式)
            end_date: 終了日 (YYYY-MM-DD形式)

        Returns:
            SoldRecordのリスト

        Raises:
            FetchCancelled: キャンセルされた場合
            ValueError: API Key/Database IDが未設定の場合
        """
        if not self.notion or not self.database_id:
            raise ValueError("API Key or Database ID not set")

        self._check_cancelled()

        all_results = []
        has_more = True
        start_cursor = None
        page_count = 0

        self._report_progress(0, -1, "売上データを取得中...")

        while has_more:
            self._check_cancelled()

            query_params = {
                "database_id": self.database_id,
                "filter": {
                    "and": [
                        {"property": "在庫状況", "status": {"equals": "売却済み"}},
                        {"property": "売却日", "date": {"on_or_after": start_date}},
                        {"property": "売却日", "date": {"before": end_date}}
                    ]
                },
                "page_size": 100
            }
            if start_cursor:
                query_params["start_cursor"] = start_cursor

            results = self._query_with_retry(query_params)
            all_results.extend(results["results"])
            has_more = results.get("has_more", False)
            start_cursor = results.get("next_cursor")

            page_count += 1
            self._report_progress(
                len(all_results), -1,
                f"売上データを取得中... ({len(all_results)}件取得済み)"
            )

        # Parse and validate
        records = []
        for page in all_results:
            flat_data = self._flatten_notion_page(page)
            try:
                record = SoldRecord(**flat_data)
                # Calculate year_month
                if record.sold_date:
                    record.sold_year_month = f"{record.sold_date.year}年{record.sold_date.month}月"
                records.append(record)
            except Exception as e:
                print(f"Skipping invalid record: {e}")
                continue
        
        return records

    def fetch_purchase_data(self, start_date: str, end_date: str) -> List[PurchaseRecord]:
        """Fetch items from Notion whose 仕入れ日 falls within the date range

        「仕入れ日」プロパティに対するサーバーサイドの date フィルタで必要な
        レコードだけを取得する。帰属月は仕入れ日ベースで計算される。
        仕入れ日が未入力のアイテムはフィルタにより集計対象外になる（仕様）。

        Args:
            start_date: 開始日 (YYYY-MM-DD形式)
            end_date: 終了日 (YYYY-MM-DD形式)

        Returns:
            PurchaseRecordのリスト

        Raises:
            FetchCancelled: キャンセルされた場合
            ValueError: API Key/Database IDが未設定の場合
        """
        if not self.notion or not self.database_id:
            raise ValueError("API Key or Database ID not set")

        self._check_cancelled()

        all_results = []
        has_more = True
        start_cursor = None
        page_count = 0

        # PurchaseRecord が必要とするプロパティのみ取得して転送量を削る。
        # ID 解決に失敗した場合は絞り込みなし（全プロパティ取得）にフォールバック。
        filter_properties = self._get_property_ids(
            ["仕入れ原価", "仕入れ先名", "仕入先カテゴリ", "作業担当", "仕入れ日"]
        )
        if not filter_properties:
            print("filter_properties のID解決に失敗しました（全プロパティ取得で続行）")
            filter_properties = None

        self._report_progress(0, -1, "仕入データを取得中...")

        while has_more:
            self._check_cancelled()

            query_params = {
                "database_id": self.database_id,
                "filter": {
                    "and": [
                        {"property": "仕入れ日", "date": {"on_or_after": start_date}},
                        {"property": "仕入れ日", "date": {"before": end_date}},
                    ]
                },
                "page_size": 100,
            }
            if start_cursor:
                query_params["start_cursor"] = start_cursor

            results = self._query_with_retry(query_params, filter_properties=filter_properties)
            all_results.extend(results["results"])
            has_more = results.get("has_more", False)
            start_cursor = results.get("next_cursor")

            page_count += 1
            self._report_progress(
                len(all_results), -1,
                f"仕入データを取得中... ({len(all_results)}件取得済み)"
            )

        records = []
        for page in all_results:
            flat_data = self._flatten_notion_page(page)
            try:
                record = PurchaseRecord(**flat_data)
                # Calculate year_month
                if record.purchase_date:
                    record.purchase_year_month = f"{record.purchase_date.year}年{record.purchase_date.month}月"
                records.append(record)
            except Exception as e:
                print(f"Skipping invalid purchase record: {e}")
                continue

        return records

    def _flatten_notion_page(self, page: Dict[str, Any]) -> Dict[str, Any]:
        """Flatten Notion page properties into a dictionary"""
        properties = page.get("properties", {})
        row = {}
        row["id"] = page.get("id")
        row["Created time"] = page.get("created_time") # Map created_time to "Created time" for schema

        for prop_name, prop_value in properties.items():
            prop_name = prop_name.strip()  # Notionプロパティ名の余分な空白を除去
            prop_type = prop_value["type"]
            
            if prop_type == "title":
                row[prop_name] = prop_value["title"][0]["plain_text"] if prop_value["title"] else ""
            elif prop_type == "rich_text":
                row[prop_name] = prop_value["rich_text"][0]["plain_text"] if prop_value["rich_text"] else ""
            elif prop_type == "number":
                row[prop_name] = prop_value.get("number", 0)
            elif prop_type == "select":
                row[prop_name] = prop_value["select"]["name"] if prop_value["select"] else ""
            elif prop_type == "status":
                row[prop_name] = prop_value["status"]["name"] if prop_value["status"] else ""
            elif prop_type == "multi_select":
                row[prop_name] = ", ".join([item["name"] for item in prop_value["multi_select"]])
            elif prop_type == "date":
                row[prop_name] = prop_value["date"]["start"] if prop_value["date"] else None
            elif prop_type == "checkbox":
                row[prop_name] = prop_value["checkbox"]
            elif prop_type == "url":
                row[prop_name] = prop_value.get("url", "")
            elif prop_type == "email":
                row[prop_name] = prop_value.get("email", "")
            elif prop_type == "phone_number":
                row[prop_name] = prop_value.get("phone_number", "")
            elif prop_type == "formula":
                formula = prop_value.get("formula", {})
                if formula.get("type") == "string":
                    row[prop_name] = formula.get("string", "")
                elif formula.get("type") == "number":
                    row[prop_name] = formula.get("number", 0)
            elif prop_type == "rollup":
                rollup = prop_value.get("rollup", {})
                rollup_type = rollup.get("type")

                if rollup_type == "array":
                    array_data = rollup.get("array", [])
                    values = []
                    for item in array_data:
                        item_type = item.get("type")
                        if item_type == "title" and item.get("title"):
                            if item["title"]:
                                values.append(item["title"][0]["plain_text"])
                        elif item_type == "rich_text" and item.get("rich_text"):
                            if item["rich_text"]:
                                values.append(item["rich_text"][0]["plain_text"])
                        elif item_type == "select" and item.get("select"):
                            values.append(item["select"]["name"])
                        elif item_type == "number":
                            values.append(str(item.get("number", "")))
                        elif item_type == "formula":
                            # ロールアップ内のformulaタイプに対応
                            formula = item.get("formula", {})
                            formula_type = formula.get("type")
                            if formula_type == "string":
                                string_val = formula.get("string", "")
                                if string_val:
                                    values.append(string_val)
                            elif formula_type == "number":
                                num_val = formula.get("number")
                                if num_val is not None:
                                    values.append(str(num_val))
                        # Add simplistic support for others
                    row[prop_name] = ", ".join(values) if values else ""
                elif rollup_type == "number":
                    row[prop_name] = rollup.get("number", 0)
                elif rollup_type == "date":
                    date_val = rollup.get("date")
                    row[prop_name] = date_val["start"] if date_val else None
            elif prop_type == "relation":
                 # Simplify rels
                 row[prop_name] = str(prop_value.get("relation", []))
            elif prop_type == "people":
                 people = prop_value.get("people", [])
                 names = [p.get("name", "") for p in people if p.get("name")]
                 row[prop_name] = ", ".join(names) if names else ""
            # Add other types as needed
        return row

    def process_pivot_data(self, sales: List[SoldRecord], purchases: List[PurchaseRecord], months: List[str]) -> Dict[str, Any]:
        """
        Process raw records into pivot tables.
        Returns dictionary containing DataFrames and mapping dicts.
        """
        if not sales and not purchases:
            return {}

        # 1. Convert to DataFrame for easier manipulation
        sales_data = [s.model_dump() for s in sales]
        df_sales = pd.DataFrame(sales_data) if sales_data else pd.DataFrame()
        
        purchase_data = [p.model_dump() for p in purchases]
        df_purchases = pd.DataFrame(purchase_data) if purchase_data else pd.DataFrame()

        # Mapping dictionaries
        company_category_map = {}
        # Purchase specific map
        purchase_company_category_map = {}

        # If we have sales data
        if not df_sales.empty:
            # Normalize missing months
            for month in months:
                if month not in df_sales['sold_year_month'].unique():
                    # Just ensure we handle missing months in pivot
                    pass

            # Update category map from sales (販売先カテゴリのみ)
            for _, row in df_sales[['sales_channel', 'sales_channel_category']].drop_duplicates().iterrows():
                 if row['sales_channel'] and row['sales_channel'] != '不明':
                    company_category_map.setdefault(row['sales_channel'], row['sales_channel_category'])

        # If we have purchase data
        if not df_purchases.empty:
            for _, row in df_purchases[['supplier', 'supplier_category']].drop_duplicates().iterrows():
                if row['supplier'] and row['supplier'] != '不明':
                    purchase_company_category_map.setdefault(row['supplier'], row['supplier_category'])

        # カテゴリ定義: 市場・業販 vs 小売り
        wholesale_categories = {'市場', '業販', 'ネット'}  # 市場・業販カテゴリ
        retail_categories = {'小売', '小売り'}  # 小売りカテゴリ

        # --- Pivot 1: Sales (販売媒体のみ) ---
        pivot_sales = pd.DataFrame()
        pivot_sales_wholesale = pd.DataFrame()  # 市場・業販
        pivot_sales_retail = pd.DataFrame()     # 小売り
        if not df_sales.empty:
            # 販売媒体の売上のみ
            pivot_sales = df_sales.pivot_table(index='sales_channel', columns='sold_year_month', values='sales_amount', aggfunc='sum', fill_value=0)

            # 市場・業販の売上
            df_wholesale = df_sales[df_sales['sales_channel_category'].isin(wholesale_categories)]
            if not df_wholesale.empty:
                pivot_sales_wholesale = df_wholesale.pivot_table(index='sales_channel', columns='sold_year_month', values='sales_amount', aggfunc='sum', fill_value=0)

            # 小売りの売上
            df_retail = df_sales[df_sales['sales_channel_category'].isin(retail_categories)]
            if not df_retail.empty:
                pivot_sales_retail = df_retail.pivot_table(index='sales_channel', columns='sold_year_month', values='sales_amount', aggfunc='sum', fill_value=0)

        # --- Pivot 2: Profit (販売媒体のみ) ---
        pivot_profit = pd.DataFrame()
        pivot_profit_wholesale = pd.DataFrame()  # 市場・業販
        pivot_profit_retail = pd.DataFrame()     # 小売り
        if not df_sales.empty:
            # 販売媒体の利益のみ
            pivot_profit = df_sales.pivot_table(index='sales_channel', columns='sold_year_month', values='profit', aggfunc='sum', fill_value=0)

            # 市場・業販の利益
            if not df_wholesale.empty:
                pivot_profit_wholesale = df_wholesale.pivot_table(index='sales_channel', columns='sold_year_month', values='profit', aggfunc='sum', fill_value=0)

            # 小売りの利益
            if not df_retail.empty:
                pivot_profit_retail = df_retail.pivot_table(index='sales_channel', columns='sold_year_month', values='profit', aggfunc='sum', fill_value=0)

        # --- Pivot 3: Purchase Cost ---
        pivot_purchase = pd.DataFrame()
        if not df_purchases.empty:
            pivot_purchase = df_purchases.pivot_table(index='supplier', columns='purchase_year_month', values='cost_price', aggfunc='sum', fill_value=0)

        # Ensure all months are present in columns
        all_pivots = [
            pivot_sales, pivot_profit, pivot_purchase,
            pivot_sales_wholesale, pivot_sales_retail,
            pivot_profit_wholesale, pivot_profit_retail
        ]
        for pivot in all_pivots:
            if not pivot.empty:
                for m in months:
                    if m not in pivot.columns:
                        pivot[m] = 0
                # Reorder columns
                pivot = pivot[months]

        # Calculate Category Aggregates
        # 動的にカテゴリを取得
        all_categories = set()
        all_categories.update(company_category_map.values())
        all_categories.update(purchase_company_category_map.values())
        all_categories.discard('不明')  # '不明'は除外
        category_list = sorted(list(all_categories)) if all_categories else ['その他']

        category_sales = self._aggregate_by_category(pivot_sales, company_category_map, months, category_list, default_cat="その他")
        category_profit = self._aggregate_by_category(pivot_profit, company_category_map, months, category_list, default_cat="その他")

        # 仕入高用のカテゴリ別集計（小売りを除外）
        purchase_category_list = sorted([c for c in category_list if c not in retail_categories])
        category_purchase = self._aggregate_by_category(pivot_purchase, purchase_company_category_map, months, purchase_category_list, default_cat="その他")

        # 市場・業販用のカテゴリ別集計（市場計、業販計、ネット計）
        wholesale_category_list = sorted([c for c in wholesale_categories if c in company_category_map.values()])
        category_sales_wholesale = self._aggregate_by_category(pivot_sales_wholesale, company_category_map, months, wholesale_category_list, default_cat=None)
        category_profit_wholesale = self._aggregate_by_category(pivot_profit_wholesale, company_category_map, months, wholesale_category_list, default_cat=None)

        return {
            'pivot_sales': pivot_sales,
            'pivot_profit': pivot_profit,
            'pivot_purchase': pivot_purchase,
            'pivot_sales_wholesale': pivot_sales_wholesale,
            'pivot_sales_retail': pivot_sales_retail,
            'pivot_profit_wholesale': pivot_profit_wholesale,
            'pivot_profit_retail': pivot_profit_retail,
            'company_category_map': company_category_map,
            'purchase_company_category_map': purchase_company_category_map,
            'category_sales': category_sales,
            'category_profit': category_profit,
            'category_purchase': category_purchase,
            'category_sales_wholesale': category_sales_wholesale,
            'category_profit_wholesale': category_profit_wholesale,
            'sales_records': df_sales,
            'purchase_records': df_purchases
        }

    def _aggregate_by_category(self, pivot: pd.DataFrame, mapping: Dict[str, str], months: List[str], categories: List[str], default_cat=None) -> Dict[str, Dict[str, float]]:
        result = {cat: {m: 0.0 for m in months} for cat in categories}
        if pivot.empty:
            return result
        
        for company in pivot.index:
            cat = mapping.get(company, default_cat)
            if cat in result:
                for m in months:
                    if m in pivot.columns:
                        result[cat][m] += pivot.loc[company, m]
        return result

    def generate_excel(self, file_path: str, data: Dict[str, Any], months: List[str]):
        """Generate Excel file from processed data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "財務集計"

        current_row = 1
        
        # Helper styles
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        category_fill = PatternFill(start_color="E0F2F7", end_color="E0F2F7", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center')

        # 1. Summary Section (Optional, but present in original)
        # Re-calc summary from raw sales records if possible
        df_sales = data.get('sales_records')
        df_purchases = data.get('purchase_records')
        has_sales = df_sales is not None and not df_sales.empty
        has_purchases = df_purchases is not None and not df_purchases.empty
        # 売上か仕入のどちらかがあれば全体合算ブロックを描画する
        if has_sales or has_purchases:
             ws.cell(row=current_row, column=1, value="全体合算").font = Font(bold=True, size=14)
             current_row += 1

             # Headers
             ws.cell(row=current_row, column=1, value="内訳").fill = header_fill
             for i, m in enumerate(months, 2):
                 c = ws.cell(row=current_row, column=i, value=m)
                 c.fill = header_fill
                 c.alignment = center_align
                 c.font = bold_font
             ws.cell(row=current_row, column=len(months)+2, value="計").fill = header_fill
             current_row += 1

             # 仕入高は仕入レコード（仕入れ日ベース）、その他は売却レコードから集計
             summary_items = ["売上", "原価", "粗利", "販売手数料", "送料", "販売利益", "仕入高"]
             # Calculate sums
             sums = {item: {m: 0.0 for m in months} for item in summary_items}

             if has_sales:
                 for _, row in df_sales.iterrows():
                     m = row['sold_year_month']
                     if m in months:
                         sums["売上"][m] += row['sales_amount']
                         sums["原価"][m] += row['cost_price']
                         sums["販売手数料"][m] += row['commission']
                         sums["送料"][m] += row['shipping_cost']
                         sums["販売利益"][m] += row['profit']
                         sums["粗利"][m] = sums["売上"][m] - sums["原価"][m]

             if has_purchases:
                 for _, row in df_purchases.iterrows():
                     m = row['purchase_year_month']
                     if m in months:
                         sums["仕入高"][m] += row['cost_price']

             for item in summary_items:
                 ws.cell(row=current_row, column=1, value=item).fill = total_fill
                 total = 0
                 for i, m in enumerate(months, 2):
                     val = sums[item][m]
                     c = ws.cell(row=current_row, column=i, value=val)
                     c.number_format = '#,##0'
                     c.fill = total_fill
                     total += val
                 ws.cell(row=current_row, column=len(months)+2, value=total).number_format = '#,##0'
                 current_row += 1
             
             current_row += 1

        # 2. Pivot Sections - 4分割（市場・業販 / 小売り）+ 仕入高
        sections = [
            ("企業別売上(市場・業販)", data.get('pivot_sales_wholesale'), data.get('category_sales_wholesale')),
            ("企業別販売利益(市場・業販)", data.get('pivot_profit_wholesale'), data.get('category_profit_wholesale')),
            ("企業別売上(小売り)", data.get('pivot_sales_retail'), None),
            ("企業別販売利益(小売り)", data.get('pivot_profit_retail'), None),
            ("企業別仕入高", data.get('pivot_purchase'), data.get('category_purchase')),
        ]

        # Assignee mapping - 販売担当者（ロールアップ）を使用してカテゴリごとに分離
        # カテゴリ定義
        wholesale_categories = {'市場', '業販', 'ネット'}  # 市場・業販カテゴリ
        retail_categories = {'小売', '小売り'}  # 小売りカテゴリ

        # 市場・業販用：担当者マッピング
        wholesale_assignee_map = {}
        # 小売り用：担当者マッピング
        retail_assignee_map = {}

        if df_sales is not None and not df_sales.empty:
            for _, r in df_sales.iterrows():
                # 販売担当者（ロールアップ）を使用
                sales_assignee = r.get('sales_assignee')
                sales_channel = r.get('sales_channel')
                category = r.get('sales_channel_category', '')
                if sales_assignee and sales_channel and sales_channel != '不明':
                    if category in wholesale_categories:
                        if sales_assignee not in wholesale_assignee_map:
                            wholesale_assignee_map[sales_assignee] = set()
                        wholesale_assignee_map[sales_assignee].add(sales_channel)
                    elif category in retail_categories:
                        if sales_assignee not in retail_assignee_map:
                            retail_assignee_map[sales_assignee] = set()
                        retail_assignee_map[sales_assignee].add(sales_channel)

        # Sort lists
        wholesale_assignee_map = {k: sorted(list(v)) for k, v in wholesale_assignee_map.items()}
        retail_assignee_map = {k: sorted(list(v)) for k, v in retail_assignee_map.items()}

        assignee_colors = [
            PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
            PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid"),
            PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
            PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
            PatternFill(start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"),
            PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),
        ]

        for title, pivot, cats in sections:
            if pivot is None or pivot.empty:
                continue

            # セクションごとに適切な担当者マップを選択
            if '市場・業販' in title:
                assignee_map = wholesale_assignee_map
            elif '小売り' in title:
                assignee_map = retail_assignee_map
            else:
                assignee_map = {}

            ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=14)
            current_row += 1

            # Headers
            ws.cell(row=current_row, column=1).fill = header_fill
            for i, m in enumerate(months, 2):
                c = ws.cell(row=current_row, column=i, value=m)
                c.fill = header_fill
                c.font = bold_font
                c.alignment = center_align
            ws.cell(row=current_row, column=len(months)+2, value="計").fill = header_fill
            current_row += 1

            # 仕入高・市場/業販セクションは企業→カテゴリ小計の並び
            if (title == "企業別仕入高" or '市場・業販' in title) and cats:
                if title == "企業別仕入高":
                    category_map_for_section = data.get('purchase_company_category_map', {})
                else:
                    category_map_for_section = data.get('company_category_map', {})
                used_companies = set()

                # カテゴリ順序を定義（市場、ネット、業販、その他）
                category_order = ['市場', 'ネット', '業販', 'その他']
                sorted_cats = sorted(cats.items(), key=lambda x: category_order.index(x[0]) if x[0] in category_order else 999)

                for cat_name, cat_vals in sorted_cats:
                    # そのカテゴリに属する企業を先に出力
                    for company in pivot.index:
                        company_cat = category_map_for_section.get(company, 'その他')
                        if company_cat == cat_name and company not in used_companies:
                            used_companies.add(company)
                            ws.cell(row=current_row, column=1, value=company)
                            row_total = 0
                            for i, m in enumerate(months, 2):
                                val = pivot.loc[company, m] if m in pivot.columns else 0
                                c = ws.cell(row=current_row, column=i, value=val)
                                c.number_format = '#,##0'
                                row_total += val
                            ws.cell(row=current_row, column=len(months)+2, value=row_total).number_format = '#,##0'
                            current_row += 1

                    # カテゴリ小計を後に出力
                    c_label = ws.cell(row=current_row, column=1, value=f"{cat_name}計")
                    c_label.fill = category_fill
                    c_label.font = bold_font
                    row_total = 0
                    for i, m in enumerate(months, 2):
                        val = cat_vals.get(m, 0)
                        c = ws.cell(row=current_row, column=i, value=val)
                        c.number_format = '#,##0'
                        c.fill = category_fill
                        c.font = bold_font
                        row_total += val
                    c_total = ws.cell(row=current_row, column=len(months)+2, value=row_total)
                    c_total.fill = category_fill
                    c_total.font = bold_font
                    c_total.number_format = '#,##0'
                    current_row += 1

                # 市場・業販の場合はユーザー計を合計の上にまとめて出力
                if '市場・業販' in title and assignee_map:
                    for assignee, companies in sorted(assignee_map.items()):
                        # 担当者の小計を計算
                        subtotal = {m: 0.0 for m in months}
                        has_data = False
                        for company in companies:
                            if company in pivot.index:
                                has_data = True
                                for m in months:
                                    if m in pivot.columns:
                                        subtotal[m] += pivot.loc[company, m]

                        if has_data:
                            color_idx = list(sorted(assignee_map.keys())).index(assignee)
                            fill = assignee_colors[color_idx % len(assignee_colors)]

                            label = f"{assignee}計"
                            c_label = ws.cell(row=current_row, column=1, value=label)
                            c_label.fill = fill
                            c_label.font = bold_font
                            grand_subtotal = 0
                            for i, m in enumerate(months, 2):
                                val = subtotal[m]
                                c = ws.cell(row=current_row, column=i, value=val)
                                c.number_format = '#,##0'
                                c.fill = fill
                                c.font = bold_font
                                grand_subtotal += val
                            c_total = ws.cell(row=current_row, column=len(months)+2, value=grand_subtotal)
                            c_total.fill = fill
                            c_total.font = bold_font
                            c_total.number_format = '#,##0'
                            current_row += 1
            else:
                # 通常のセクション（担当者別グループ）
                used_companies = set()
                if assignee_map:
                    color_idx = 0
                    for assignee, companies in sorted(assignee_map.items()):
                        fill = assignee_colors[color_idx % len(assignee_colors)]
                        color_idx += 1

                        subtotal = {m: 0.0 for m in months}

                        has_rows = False
                        for company in companies:
                            if company in pivot.index:
                                has_rows = True
                                used_companies.add(company)
                                ws.cell(row=current_row, column=1, value=company)
                                row_total = 0
                                for i, m in enumerate(months, 2):
                                    val = pivot.loc[company, m] if m in pivot.columns else 0
                                    c = ws.cell(row=current_row, column=i, value=val)
                                    c.number_format = '#,##0'
                                    row_total += val
                                    subtotal[m] += val

                                ws.cell(row=current_row, column=len(months)+2, value=row_total).number_format = '#,##0'
                                current_row += 1

                        if has_rows:
                            # Subtotal row (ユーザー別小計)
                            label = f"{assignee}計"
                            c_label = ws.cell(row=current_row, column=1, value=label)
                            c_label.fill = fill
                            c_label.font = bold_font
                            grand_subtotal = 0
                            for i, m in enumerate(months, 2):
                                val = subtotal[m]
                                c = ws.cell(row=current_row, column=i, value=val)
                                c.number_format = '#,##0'
                                c.fill = fill
                                c.font = bold_font
                                grand_subtotal += val
                            c_total = ws.cell(row=current_row, column=len(months)+2, value=grand_subtotal)
                            c_total.fill = fill
                            c_total.font = bold_font
                            c_total.number_format = '#,##0'
                            current_row += 1

                # Remaining companies
                remaining = [c for c in pivot.index if c not in used_companies]
                for company in remaining:
                    ws.cell(row=current_row, column=1, value=company)
                    row_total = 0
                    for i, m in enumerate(months, 2):
                        val = pivot.loc[company, m] if m in pivot.columns else 0
                        c = ws.cell(row=current_row, column=i, value=val)
                        c.number_format = '#,##0'
                        row_total += val
                    ws.cell(row=current_row, column=len(months)+2, value=row_total).number_format = '#,##0'
                    current_row += 1

                # Category Totals (カテゴリ別小計) - 仕入高以外
                if cats:
                    for cat_name, cat_vals in sorted(cats.items()):
                        c_label = ws.cell(row=current_row, column=1, value=f"{cat_name}計")
                        c_label.fill = category_fill
                        c_label.font = bold_font
                        row_total = 0
                        for i, m in enumerate(months, 2):
                            val = cat_vals.get(m, 0)
                            c = ws.cell(row=current_row, column=i, value=val)
                            c.number_format = '#,##0'
                            c.fill = category_fill
                            c.font = bold_font
                            row_total += val
                        c_total = ws.cell(row=current_row, column=len(months)+2, value=row_total)
                        c_total.fill = category_fill
                        c_total.font = bold_font
                        c_total.number_format = '#,##0'
                        current_row += 1

            # Grand Total (総計)
            c_label = ws.cell(row=current_row, column=1, value="合計")
            c_label.fill = total_fill
            c_label.font = bold_font
            grand_row_total = 0
            for i, m in enumerate(months, 2):
                col_sum = pivot[m].sum() if m in pivot.columns else 0
                c = ws.cell(row=current_row, column=i, value=col_sum)
                c.number_format = '#,##0'
                c.fill = total_fill
                c.font = bold_font
                grand_row_total += col_sum
            c_total = ws.cell(row=current_row, column=len(months)+2, value=grand_row_total)
            c_total.fill = total_fill
            c_total.font = bold_font
            c_total.number_format = '#,##0'
            current_row += 1
            
            current_row += 1 # Space
        
        # Auto-width
        ws.column_dimensions['A'].width = 25
        for i in range(2, len(months)+4):
            col_letter = chr(64 + i) if i <= 26 else chr(64 + (i-1)//26) + chr(65 + (i-1)%26)
            ws.column_dimensions[col_letter].width = 12

        wb.save(file_path)

    def fetch_daily_sales_data(self, start_date: str, end_date: str) -> List[DailySoldRecord]:
        """日別出力用: 売却日が期間内 かつ 在庫状況=売却済み のレコードを取得

        Args:
            start_date: 開始日 (YYYY-MM-DD形式, inclusive)
            end_date: 終了日の翌日 (YYYY-MM-DD形式, exclusive)

        Returns:
            DailySoldRecordのリスト（作成日時昇順）
        """
        if not self.notion or not self.database_id:
            raise ValueError("API Key or Database ID not set")

        self._check_cancelled()

        all_results = []
        has_more = True
        start_cursor = None

        self._report_progress(0, -1, "売上データを取得中...")

        while has_more:
            self._check_cancelled()

            query_params = {
                "database_id": self.database_id,
                "filter": {
                    "and": [
                        {"property": "在庫状況", "status": {"equals": "売却済み"}},
                        {"property": "売却日", "date": {"on_or_after": start_date}},
                        {"property": "売却日", "date": {"before": end_date}},
                    ]
                },
                "sorts": [{"timestamp": "created_time", "direction": "ascending"}],
                "page_size": 100,
            }
            if start_cursor:
                query_params["start_cursor"] = start_cursor

            results = self._query_with_retry(query_params)
            all_results.extend(results["results"])
            has_more = results.get("has_more", False)
            start_cursor = results.get("next_cursor")

            self._report_progress(
                len(all_results), -1,
                f"売上データを取得中... ({len(all_results)}件取得済み)"
            )

        print(f"[DEBUG] fetch_daily_sales_data: API returned {len(all_results)} raw records")

        records = []
        for page in all_results:
            flat_data = self._flatten_notion_page(page)
            try:
                record = DailySoldRecord(**flat_data)
                records.append(record)
            except Exception as e:
                print(f"[DEBUG] Skipping daily sold record: {e}")
                print(f"[DEBUG]   flat_data keys: {list(flat_data.keys())}")
                continue

        print(f"[DEBUG] fetch_daily_sales_data: parsed {len(records)} records")
        return records

    def fetch_daily_purchase_data(self, start_date: str, end_date: str) -> List[DailyPurchaseRecord]:
        """日別出力用: 仕入れ日が期間内 の全ステータスレコードを取得

        Args:
            start_date: 開始日 (YYYY-MM-DD形式, inclusive)
            end_date: 終了日の翌日 (YYYY-MM-DD形式, exclusive)

        Returns:
            DailyPurchaseRecordのリスト（作成日時昇順）
        """
        if not self.notion or not self.database_id:
            raise ValueError("API Key or Database ID not set")

        self._check_cancelled()

        all_results = []
        has_more = True
        start_cursor = None

        self._report_progress(0, -1, "仕入れデータを取得中...")

        while has_more:
            self._check_cancelled()

            query_params = {
                "database_id": self.database_id,
                "filter": {
                    "and": [
                        {"property": "仕入れ日", "date": {"on_or_after": start_date}},
                        {"property": "仕入れ日", "date": {"before": end_date}},
                    ]
                },
                "sorts": [{"timestamp": "created_time", "direction": "ascending"}],
                "page_size": 100,
            }
            if start_cursor:
                query_params["start_cursor"] = start_cursor

            results = self._query_with_retry(query_params)
            all_results.extend(results["results"])
            has_more = results.get("has_more", False)
            start_cursor = results.get("next_cursor")

            self._report_progress(
                len(all_results), -1,
                f"仕入れデータを取得中... ({len(all_results)}件取得済み)"
            )

        records = []
        for page in all_results:
            flat_data = self._flatten_notion_page(page)
            try:
                record = DailyPurchaseRecord(**flat_data)
                records.append(record)
            except Exception as e:
                print(f"Skipping invalid daily purchase record: {e}")
                continue

        return records

    def generate_daily_excel(
        self,
        file_path: str,
        sales: List[DailySoldRecord],
        purchases: List[DailyPurchaseRecord],
    ):
        """日別出力: 売上セクション + 仕入れセクションをシート別に出力"""
        wb = Workbook()

        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")

        # ---- 売上シート ----
        ws_sales = wb.active
        ws_sales.title = "売上"

        sales_columns = [
            ("商品名",       lambda r: r.product_name),
            ("型番",         lambda r: r.model_number),
            ("製番",         lambda r: r.serial_number),
            ("売却日",       lambda r: r.sold_date),
            ("売上金",       lambda r: r.sales_amount),
            ("仕入れ金",     lambda r: r.purchase_cost),
            ("仕入れ手数料", lambda r: r.purchase_fee),
            ("仕入れ原価",   lambda r: r.cost_price),
            ("送料",         lambda r: r.shipping_cost),
            ("送料計算方法", lambda r: r.shipping_method),
            ("販売手数料",   lambda r: r.commission),
            ("純利益",       lambda r: r.profit),
            ("利益率",       lambda r: r.profit_rate),
            ("仕入れ先",     lambda r: r.supplier),
            ("仕入れ日",     lambda r: r.purchase_date),
            ("販売媒体名",   lambda r: r.sales_channel),
            ("作業担当",     lambda r: r.assignee),
            ("伝票番号",     lambda r: r.slip_number),
            ("発送伝票番号", lambda r: r.shipping_slip_number),
            ("購入者名",     lambda r: r.buyer_name),
        ]
        numeric_sales_cols = {
            "売上金", "仕入れ金", "仕入れ手数料", "仕入れ原価",
            "送料", "販売手数料", "純利益", "発送伝票番号",
        }

        # ヘッダー行
        for col_idx, (col_name, _) in enumerate(sales_columns, 1):
            cell = ws_sales.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_align

        # データ行
        for row_idx, record in enumerate(sales, 2):
            for col_idx, (col_name, getter) in enumerate(sales_columns, 1):
                value = getter(record)
                cell = ws_sales.cell(row=row_idx, column=col_idx, value=value)
                if col_name in numeric_sales_cols and value is not None:
                    cell.number_format = "#,##0"

        # 合計行
        total_row = len(sales) + 2
        numeric_totals = {
            "売上金", "仕入れ金", "仕入れ手数料", "仕入れ原価",
            "送料", "販売手数料", "純利益",
        }
        # 利益率の平均（None を除いた件数で割る）
        profit_rates = [r.profit_rate for r in sales if r.profit_rate is not None]
        avg_profit_rate = sum(profit_rates) / len(profit_rates) if profit_rates else None

        for col_idx, (col_name, getter) in enumerate(sales_columns, 1):
            if col_name in numeric_totals:
                total = sum(getter(r) or 0 for r in sales)
                cell = ws_sales.cell(row=total_row, column=col_idx, value=total)
                cell.number_format = "#,##0"
                cell.fill = total_fill
                cell.font = bold_font
            elif col_name == "利益率":
                cell = ws_sales.cell(row=total_row, column=col_idx, value=avg_profit_rate)
                cell.fill = total_fill
                cell.font = bold_font
            elif col_idx == 1:
                cell = ws_sales.cell(row=total_row, column=col_idx, value="合計")
                cell.fill = total_fill
                cell.font = bold_font

        # 列幅
        ws_sales.column_dimensions["A"].width = 28
        for col_idx, (col_name, _) in enumerate(sales_columns, 1):
            if col_idx > 1:
                col_letter = ws_sales.cell(row=1, column=col_idx).column_letter
                ws_sales.column_dimensions[col_letter].width = 14

        # ---- 仕入れシート ----
        ws_purchase = wb.create_sheet(title="仕入れ")

        purchase_columns = [
            ("商品名",         lambda r: r.product_name),
            ("型番",           lambda r: r.model_number),
            ("製番",           lambda r: r.serial_number),
            ("メーカー",       lambda r: r.maker),
            ("カテゴリー",     lambda r: r.category),
            ("サイズ",         lambda r: r.size),
            ("年式",           lambda r: r.year),
            ("ランク",         lambda r: r.rank),
            ("仕入れ日",       lambda r: r.purchase_date),
            ("仕入れ金",       lambda r: r.purchase_cost),
            ("仕入れ手数料",   lambda r: r.purchase_fee),
            ("仕入れ原価",     lambda r: r.cost_price),
            ("仕入れ先",       lambda r: r.supplier),
            ("仕入先カテゴリ", lambda r: r.supplier_category),
            ("在庫状況",       lambda r: r.stock_status),
            ("作業担当",       lambda r: r.assignee),
            ("見込み売上",     lambda r: r.estimated_sales),
        ]
        numeric_purchase_cols = {
            "仕入れ金", "仕入れ手数料", "仕入れ原価", "見込み売上",
        }

        # ヘッダー行
        for col_idx, (col_name, _) in enumerate(purchase_columns, 1):
            cell = ws_purchase.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_align

        # データ行
        for row_idx, record in enumerate(purchases, 2):
            for col_idx, (col_name, getter) in enumerate(purchase_columns, 1):
                value = getter(record)
                cell = ws_purchase.cell(row=row_idx, column=col_idx, value=value)
                if col_name in numeric_purchase_cols and value is not None:
                    cell.number_format = "#,##0"

        # 合計行
        total_row_p = len(purchases) + 2
        numeric_purchase_totals = {"仕入れ金", "仕入れ手数料", "仕入れ原価", "見込み売上"}
        for col_idx, (col_name, getter) in enumerate(purchase_columns, 1):
            if col_name in numeric_purchase_totals:
                total = sum(getter(r) or 0 for r in purchases)
                cell = ws_purchase.cell(row=total_row_p, column=col_idx, value=total)
                cell.number_format = "#,##0"
                cell.fill = total_fill
                cell.font = bold_font
            elif col_idx == 1:
                cell = ws_purchase.cell(row=total_row_p, column=col_idx, value="合計")
                cell.fill = total_fill
                cell.font = bold_font

        # 列幅
        ws_purchase.column_dimensions["A"].width = 28
        for col_idx, (col_name, _) in enumerate(purchase_columns, 1):
            if col_idx > 1:
                col_letter = ws_purchase.cell(row=1, column=col_idx).column_letter
                ws_purchase.column_dimensions[col_letter].width = 14

        wb.save(file_path)
