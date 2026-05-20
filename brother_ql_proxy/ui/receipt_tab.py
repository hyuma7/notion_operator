"""
領収書タブ
- Notionから直近30件取得 or 商品名で検索
- チェックボックスで複数選択
- 選択した商品の領収書プレビューを表示
- Excel形式で領収書を出力
"""

import os
import threading
from datetime import date
import flet as ft

from notion.fetch_page import fetch_recent_items, search_items, fetch_all_properties


def _fmt_currency(value) -> str:
    try:
        n = int(value)
        return f"¥{n:,}"
    except (TypeError, ValueError):
        return "¥0"


def _generate_receipt_excel(path: str, items: list[tuple[str, int]], issue_date: str):
    """複数商品の領収書Excelを生成する"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "領収書"

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 20

    thin = Side(style="thin")
    medium = Side(style="medium")
    thick = Side(style="thick")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_border = Border(left=thin, right=thin, top=thin, bottom=medium)
    hanko_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # タイトル
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "領　収　書"
    c.font = Font(name="MS Gothic", size=24, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # 発行日
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = f"発行日: {issue_date}"
    c.font = Font(name="MS Gothic", size=11)
    c.alignment = Alignment(horizontal="right", vertical="center")

    # ヘッダー行
    header_row = 4
    ws.row_dimensions[header_row].height = 22
    for col, label in [(1, "No."), (2, "品名"), (3, "数量"), (4, "金額")]:
        c = ws.cell(row=header_row, column=col)
        c.value = label
        c.font = Font(name="MS Gothic", size=11, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_border
        c.fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    # 商品行
    data_start = header_row + 1
    for idx, (item_name, amount) in enumerate(items):
        r = data_start + idx
        ws.row_dimensions[r].height = 20
        for col, val, fmt, align in [
            (1, idx + 1, None, "center"),
            (2, item_name, None, "left"),
            (3, 1, None, "center"),
            (4, amount, '"¥"#,##0', "right"),
        ]:
            c = ws.cell(row=r, column=col)
            c.value = val
            c.font = Font(name="MS Gothic", size=11)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = data_border
            if fmt:
                c.number_format = fmt

    # 合計ブロック
    subtotal = sum(a for _, a in items)
    tax = int(subtotal * 0.1)
    total = subtotal + tax

    summary_start = data_start + len(items) + 2
    for i, (label, value, bold) in enumerate([
        ("小計", subtotal, False),
        ("消費税(10%)", tax, False),
        ("合計", total, True),
    ]):
        r = summary_start + i
        ws.row_dimensions[r].height = 22
        lc = ws.cell(row=r, column=3)
        lc.value = label
        lc.font = Font(name="MS Gothic", size=11, bold=bold)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border = total_border if bold else header_border

        vc = ws.cell(row=r, column=4)
        vc.value = value
        vc.font = Font(name="MS Gothic", size=12 if bold else 11, bold=bold)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border = total_border if bold else header_border
        vc.number_format = '"¥"#,##0'

    # 署名・ハンコ
    sign_start = summary_start + 5
    for i, text in enumerate(["株式会社アーネスト", "代表取締役  斉藤 潤"]):
        r = sign_start + i
        ws.row_dimensions[r].height = 22
        c = ws.cell(row=r, column=1)
        c.value = text
        c.font = Font(name="MS Gothic", size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")

    hanko_r1 = sign_start
    hanko_r2 = sign_start + 2
    ws.merge_cells(f"D{hanko_r1}:E{hanko_r2}")
    hc = ws.cell(row=hanko_r1, column=4)
    hc.value = "株式会社アーネスト\n代表取締役\n斉藤 潤"
    hc.font = Font(name="MS Gothic", size=10, bold=True, color="8B0000")
    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    hc.border = hanko_border
    hc.fill = PatternFill(fill_type="solid", fgColor="FFE4E1")
    for r in range(hanko_r1, hanko_r2 + 1):
        ws.row_dimensions[r].height = 22

    wb.save(path)


class ReceiptTab:
    def __init__(self, proxy, page: ft.Page):
        self.proxy = proxy
        self.page = page
        self._is_loading = False
        self._selected_items: dict[str, str] = {}   # page_id → title
        self._item_data_cache: dict[str, dict] = {}  # page_id → properties data
        self._current_list: list[dict] = []
        self._build_ui()

    # ─────────────────────────────────────────────────────────────────
    # UI構築
    # ─────────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.list_column = ft.Column(
            controls=[],
            scroll=ft.ScrollMode.AUTO,
            spacing=2,
        )
        self.refresh_btn = ft.ElevatedButton(
            "直近30件",
            icon=ft.Icons.REFRESH,
            on_click=self.on_refresh,
            bgcolor=ft.Colors.BLUE_GREY_700,
            color=ft.Colors.WHITE,
        )
        self.list_status = ft.Text("", size=12, color=ft.Colors.GREY_600)

        self.search_field = ft.TextField(
            label="商品名で検索",
            hint_text="例: MacBook",
            expand=True,
            dense=True,
            on_submit=self.on_search,
        )
        self.search_btn = ft.IconButton(
            icon=ft.Icons.SEARCH,
            tooltip="検索",
            on_click=self.on_search,
        )

        self.selection_label = ft.Text(
            "0件選択中",
            size=13,
            color=ft.Colors.GREY_600,
        )
        self.clear_btn = ft.TextButton(
            "選択解除",
            on_click=self._clear_selection,
        )

        self.preview_column = ft.Column(controls=[], spacing=4, visible=False)

        self.preview_btn = ft.ElevatedButton(
            "プレビュー",
            icon=ft.Icons.PREVIEW,
            on_click=self.on_preview,
            bgcolor=ft.Colors.BLUE_700,
            color=ft.Colors.WHITE,
            disabled=True,
        )
        self.export_btn = ft.ElevatedButton(
            "Excel出力",
            icon=ft.Icons.TABLE_VIEW,
            on_click=self.on_export,
            bgcolor=ft.Colors.GREEN_700,
            color=ft.Colors.WHITE,
            disabled=True,
        )
        self.export_status = ft.Text("", size=13)

        self.content = ft.Column(
            controls=[
                ft.Text("領収書", size=18, weight=ft.FontWeight.BOLD),
                ft.Divider(height=6),

                ft.Row([
                    ft.Text("最近の更新", size=14, weight=ft.FontWeight.BOLD),
                    self.refresh_btn,
                    self.list_status,
                ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Container(
                    content=self.list_column,
                    border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
                    border_radius=6,
                    padding=ft.padding.all(6),
                    height=300,
                ),

                ft.Row([self.search_field, self.search_btn], spacing=4),
                ft.Divider(height=6),

                ft.Row([
                    self.selection_label,
                    self.clear_btn,
                ], spacing=10, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Divider(height=6),

                self.preview_column,

                ft.Row(
                    [self.preview_btn, self.export_btn, self.export_status],
                    spacing=12,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                ),
            ],
            scroll=ft.ScrollMode.AUTO,
            expand=True,
            spacing=8,
        )

    # ─────────────────────────────────────────────────────────────────
    # 一覧取得
    # ─────────────────────────────────────────────────────────────────

    def on_refresh(self, e):
        db_id = self.proxy.config.get("notion_database_id", "")
        if not db_id:
            self._snack("Notion Database ID が未設定です（設定タブで設定してください）", ft.Colors.RED)
            return
        if self._is_loading:
            return
        self._start_list_load(lambda: fetch_recent_items(db_id, limit=30))

    def on_search(self, e):
        query = self.search_field.value.strip()
        if not query:
            self._snack("検索ワードを入力してください", ft.Colors.ORANGE)
            return
        db_id = self.proxy.config.get("notion_database_id", "")
        if not db_id:
            self._snack("Notion Database ID が未設定です（設定タブで設定してください）", ft.Colors.RED)
            return
        if self._is_loading:
            return
        self._start_list_load(lambda: search_items(db_id, query))

    def _start_list_load(self, fetch_fn):
        self._is_loading = True
        self.refresh_btn.disabled = True
        self.list_status.value = "取得中..."
        self.list_column.controls.clear()
        self.page.update()

        def do():
            try:
                items = fetch_fn()

                def on_ok():
                    self._current_list = items
                    self._render_list(items)
                    self.list_status.value = f"{len(items)}件"
                    self.page.update()

                self.page.run_thread(on_ok)
            except Exception as ex:
                err = ex

                def on_err():
                    self.list_status.value = f"エラー: {err}"
                    self.page.update()

                self.page.run_thread(on_err)
            finally:
                def on_fin():
                    self._is_loading = False
                    self.refresh_btn.disabled = False
                    self.page.update()

                self.page.run_thread(on_fin)

        threading.Thread(target=do, daemon=True).start()

    def _render_list(self, items: list[dict]):
        self.list_column.controls.clear()
        for item in items:
            pid = item["page_id"]
            title = item["title"]
            dt = item["last_edited_time"][:16].replace("T", " ") if item.get("last_edited_time") else ""
            is_checked = pid in self._selected_items

            cb = ft.Checkbox(
                value=is_checked,
                on_change=lambda e, p=pid, t=title: self._on_check(e, p, t),
            )
            row = ft.Container(
                content=ft.Row([
                    cb,
                    ft.Text(dt, size=11, color=ft.Colors.GREY_600, width=115, no_wrap=True),
                    ft.Text(title, size=12, expand=True, no_wrap=True, overflow=ft.TextOverflow.ELLIPSIS),
                ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                padding=ft.padding.symmetric(horizontal=4, vertical=2),
                border_radius=4,
                bgcolor=ft.Colors.GREEN_50 if is_checked else None,
            )
            self.list_column.controls.append(row)

    def _on_check(self, e, page_id: str, title: str):
        if e.control.value:
            self._selected_items[page_id] = title
        else:
            self._selected_items.pop(page_id, None)
        self._refresh_list_colors()
        self._update_selection_state()

    def _refresh_list_colors(self):
        for ctrl in self.list_column.controls:
            row = ctrl.content
            cb = row.controls[0]
            pid = None
            for item in self._current_list:
                if item["title"] == row.controls[2].value:
                    pid = item["page_id"]
                    break
            is_checked = cb.value
            ctrl.bgcolor = ft.Colors.GREEN_50 if is_checked else None
        self.page.update()

    def _update_selection_state(self):
        n = len(self._selected_items)
        self.selection_label.value = f"{n}件選択中"
        self.selection_label.color = ft.Colors.GREEN_800 if n > 0 else ft.Colors.GREY_600
        self.preview_btn.disabled = n == 0
        self.export_btn.disabled = n == 0
        self.page.update()

    def _clear_selection(self, e=None):
        self._selected_items.clear()
        self._render_list(self._current_list)
        self._update_selection_state()
        self.preview_column.visible = False
        self.preview_column.controls.clear()
        self.export_status.value = ""
        self.page.update()

    # ─────────────────────────────────────────────────────────────────
    # 選択データ取得
    # ─────────────────────────────────────────────────────────────────

    def _fetch_selected_data(self) -> list[tuple[str, dict]]:
        """選択中の全アイテムのプロパティを取得（キャッシュ利用）"""
        result = []
        for pid, title in self._selected_items.items():
            if pid not in self._item_data_cache:
                data = fetch_all_properties(pid)
                self._item_data_cache[pid] = data
            result.append((pid, self._item_data_cache[pid]))
        return result

    # ─────────────────────────────────────────────────────────────────
    # プレビュー
    # ─────────────────────────────────────────────────────────────────

    def on_preview(self, e):
        if not self._selected_items or self._is_loading:
            return
        self._is_loading = True
        self.preview_btn.disabled = True
        self.export_btn.disabled = True
        self.export_status.value = "データ取得中..."
        self.page.update()

        def do():
            try:
                fetched = self._fetch_selected_data()

                def on_ok():
                    self._render_preview(fetched)
                    self.export_btn.disabled = False
                    self.export_status.value = ""
                    self.page.update()

                self.page.run_thread(on_ok)
            except Exception as ex:
                err = ex

                def on_err():
                    self.export_status.value = f"エラー: {err}"
                    self.export_status.color = ft.Colors.RED
                    self.page.update()

                self.page.run_thread(on_err)
            finally:
                def on_fin():
                    self._is_loading = False
                    self.preview_btn.disabled = False
                    self.page.update()

                self.page.run_thread(on_fin)

        threading.Thread(target=do, daemon=True).start()

    def _render_preview(self, fetched: list[tuple[str, dict]]):
        header_style = ft.TextStyle(size=11, weight=ft.FontWeight.BOLD)

        def cell(text, bold=False, right=False):
            return ft.DataCell(ft.Text(
                text,
                size=11,
                weight=ft.FontWeight.BOLD if bold else ft.FontWeight.NORMAL,
                text_align=ft.TextAlign.RIGHT if right else ft.TextAlign.LEFT,
            ))

        rows = []
        subtotal = 0
        for idx, (pid, data) in enumerate(fetched):
            props = data.get("properties", {})
            name_info = props.get("商品名")
            item_name = str(name_info.get("value") or self._selected_items.get(pid, "")) if name_info else self._selected_items.get(pid, "")
            amount_info = props.get("売上金")
            try:
                amount = int(amount_info.get("value") or 0) if amount_info else 0
            except (TypeError, ValueError):
                amount = 0
            subtotal += amount
            rows.append(ft.DataRow(cells=[
                cell(str(idx + 1)),
                cell(item_name),
                cell("1", right=True),
                cell(_fmt_currency(amount), right=True),
            ]))

        tax = int(subtotal * 0.1)
        total = subtotal + tax
        today_str = date.today().strftime("%Y年%m月%d日")

        item_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("No.", style=header_style)),
                ft.DataColumn(ft.Text("品名", style=header_style)),
                ft.DataColumn(ft.Text("数量", style=header_style), numeric=True),
                ft.DataColumn(ft.Text("金額", style=header_style), numeric=True),
            ],
            rows=rows,
            border=ft.border.all(1, ft.Colors.GREY_400),
            border_radius=4,
            column_spacing=20,
        )

        summary_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("")),
                ft.DataColumn(ft.Text(""), numeric=True),
            ],
            rows=[
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text("小計", size=11)),
                    ft.DataCell(ft.Text(_fmt_currency(subtotal), size=11, text_align=ft.TextAlign.RIGHT)),
                ]),
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text("消費税 (税率10%)", size=11)),
                    ft.DataCell(ft.Text(_fmt_currency(tax), size=11, text_align=ft.TextAlign.RIGHT)),
                ]),
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text("合計", size=13, weight=ft.FontWeight.BOLD)),
                    ft.DataCell(ft.Text(_fmt_currency(total), size=13, weight=ft.FontWeight.BOLD,
                                       text_align=ft.TextAlign.RIGHT)),
                ]),
            ],
            border=ft.border.all(1, ft.Colors.GREY_400),
            border_radius=4,
            column_spacing=20,
        )

        hanko = ft.Container(
            content=ft.Text(
                "株式会社アーネスト\n代表取締役\n斉藤 潤",
                size=10,
                weight=ft.FontWeight.BOLD,
                color=ft.Colors.RED_900,
                text_align=ft.TextAlign.CENTER,
            ),
            width=110,
            height=80,
            border=ft.border.all(3, ft.Colors.RED_900),
            border_radius=4,
            bgcolor=ft.Colors.with_opacity(0.15, ft.Colors.RED_200),
            alignment=ft.alignment.center,
            padding=ft.padding.all(6),
        )

        self.preview_column.controls = [
            ft.Container(
                content=ft.Column([
                    ft.Text("領　収　書", size=22, weight=ft.FontWeight.BOLD,
                            text_align=ft.TextAlign.CENTER),
                    ft.Text(f"発行日: {today_str}", size=11, text_align=ft.TextAlign.RIGHT),
                    ft.Divider(height=4),
                    item_table,
                    ft.Divider(height=4),
                    ft.Row([ft.Container(expand=True), summary_table]),
                    ft.Divider(height=8),
                    ft.Row([
                        ft.Column([
                            ft.Text("株式会社アーネスト", size=11),
                            ft.Text("代表取締役  斉藤 潤", size=11),
                        ], spacing=2),
                        ft.Container(expand=True),
                        hanko,
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ], spacing=8),
                padding=ft.padding.all(16),
                border=ft.border.all(1, ft.Colors.GREY_300),
                border_radius=8,
                bgcolor=ft.Colors.WHITE,
            )
        ]
        self.preview_column.visible = True

    # ─────────────────────────────────────────────────────────────────
    # Excel出力
    # ─────────────────────────────────────────────────────────────────

    def on_export(self, e):
        if not self._selected_items:
            return

        today_str = date.today().strftime("%Y年%m月%d日")
        fname = f"領収書_{date.today().strftime('%Y%m%d')}.xlsx"

        def save_file(ev: ft.FilePickerResultEvent):
            if not ev.path:
                return

            def do():
                try:
                    fetched = self._fetch_selected_data()
                    items_data = []
                    for pid, data in fetched:
                        props = data.get("properties", {})
                        name_info = props.get("商品名")
                        item_name = str(name_info.get("value") or self._selected_items.get(pid, "")) if name_info else self._selected_items.get(pid, "")
                        amount_info = props.get("売上金")
                        try:
                            amount = int(amount_info.get("value") or 0) if amount_info else 0
                        except (TypeError, ValueError):
                            amount = 0
                        items_data.append((item_name, amount))

                    _generate_receipt_excel(ev.path, items_data, today_str)

                    def on_ok():
                        self.export_status.value = f"保存しました: {ev.path}"
                        self.export_status.color = ft.Colors.GREEN
                        self.page.update()

                    self.page.run_thread(on_ok)
                except Exception as ex:
                    err = ex

                    def on_err():
                        self.export_status.value = f"エラー: {err}"
                        self.export_status.color = ft.Colors.RED
                        self.page.update()

                    self.page.run_thread(on_err)

            threading.Thread(target=do, daemon=True).start()

        file_picker = ft.FilePicker(on_result=save_file)
        self.page.overlay.append(file_picker)
        self.page.update()
        file_picker.save_file(
            dialog_title="領収書Excelを保存",
            file_name=fname,
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx"],
        )

    # ─────────────────────────────────────────────────────────────────
    # ユーティリティ
    # ─────────────────────────────────────────────────────────────────

    def _snack(self, msg: str, color):
        snack = ft.SnackBar(content=ft.Text(msg), bgcolor=color)
        self.page.snack_bar = snack
        self.page.snack_bar.open = True
        self.page.update()

    def create_tab(self) -> ft.Tab:
        return ft.Tab(
            text="領収書",
            icon=ft.Icons.RECEIPT,
            content=ft.Container(
                padding=ft.padding.all(20),
                content=self.content,
            ),
        )


def create_receipt_tab(proxy, page: ft.Page) -> tuple[ft.Tab, "ReceiptTab"]:
    tab = ReceiptTab(proxy, page)
    return tab.create_tab(), tab
